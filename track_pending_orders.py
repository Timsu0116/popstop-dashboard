"""
track_pending_orders.py
Detects "ordered but not yet invoiced" IKON SKUs by comparing:
  1. Weekly order emails YOU sent to Kelley Brennan (e.g. "IKON WK23&24 Weekly order")
     -> parses the attached Excel (Air Cargo + SEA sheets)
  2. shipment_db.json invoices received AFTER each order's send date

Any SKU from an order that hasn't accumulated enough quantity in invoices
received after the order date is flagged as PENDING.

Usage:
    python track_pending_orders.py
    python track_pending_orders.py --days-back 60
"""

import json, re, os, base64, urllib.parse, urllib.request, argparse
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict
import openpyxl
import io

# ─── CONFIG ────────────────────────────────────────────────────────────────────
CLIENT_ID      = "330ca31b-91e5-40df-8179-b4d842f41cbf"
TOKEN_URL_BASE = "https://login.microsoftonline.com/75a2db62-9685-4cc7-bfc0-70d615d24421/oauth2/v2.0"
SCOPES         = "https://graph.microsoft.com/Mail.Read offline_access"
TOKEN_FILE     = "ikon_token.json"
GRAPH_BASE     = "https://graph.microsoft.com/v1.0"

DB_FILE        = r"D:\PopStop_App\popstop-dashboard\data\shipment_db.json"
STREAMLIT_DATA_DIR = r"D:\PopStop_App\popstop-dashboard\data"
OUTPUT_FILE    = "pending_orders.json"

OVERDUE_DAYS   = 7   # highlight if order sent more than this many days ago and still pending
SUBJECT_HINTS  = ["weekly order", "ikon wk"]   # case-insensitive subject match


# ─── AUTH ──────────────────────────────────────────────────────────────────────
def do_refresh(token_data):
    data = urllib.parse.urlencode({
        "client_id": CLIENT_ID, "grant_type": "refresh_token",
        "refresh_token": token_data.get("refresh_token", ""), "scope": SCOPES,
    }).encode()
    req = urllib.request.Request(f"{TOKEN_URL_BASE}/token", data=data, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    with urllib.request.urlopen(req) as r:
        t = json.loads(r.read())
    with open(TOKEN_FILE, "w") as f:
        json.dump(t, f)
    return t.get("access_token")


def graph_get(token, url):
    req = urllib.request.Request(url)
    req.add_header("Authorization", f"Bearer {token}")
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())


# ─── PARSE ORDER EXCEL ─────────────────────────────────────────────────────────
def parse_order_excel(excel_bytes):
    """
    Parse the weekly order workbook (sheets: 'Air Cargo', 'SEA').
    Returns {sku: {"air": qty, "sea": qty, "total": qty, "description": str}}.

    IMPORTANT: the true ordered quantity is NOT the "Quantity" column (C) —
    that's the supplier's original quote/suggestion, which you then adjust
    up or down per-store based on the latest sales data. The authoritative
    number is the sum of the per-store allocation columns PLUS the warehouse
    column (Dunedin..Te Rapa + WAREHOUSE), which is exactly what the sheet's
    own "Total" column represents — and we keep Air Cargo vs SEA separate so
    shipping method can be tracked per SKU.

    We also capture the "Description" column. Brand-new SKUs that haven't
    been invoiced/received yet won't exist in output_company.parquet, so the
    order sheet's own description is the only product name available until
    the item shows up in the system — this lets pending orders be readable
    even for not-yet-onboarded products.

    Header row format: Code, Description, Quantity, Price, Total,
                        Still to allocate, Notes, [Dunedin..Te Rapa],
                        WAREHOUSE, Total
    """
    qty_by_sku = defaultdict(lambda: {"air": 0, "sea": 0, "description": ""})
    STORE_AND_WAREHOUSE_COLS = [
        "Dunedin", "Papanui", "Queensgate", "Riccarton",
        "Richmond", "Sylvia Park", "Te Rapa", "WAREHOUSE",
    ]

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            # Classify shipping method. Prefer scanning the sheet's own content
            # (e.g. a banner row reading "SEA Freight Order") over the sheet's
            # tab name, since some order workbooks use a generic tab name like
            # "Ikon Order" for every shipment instead of naming the tab itself
            # "Air Cargo" / "SEA".
            sheet_key = None
            scan_text = " ".join(
                str(v) for row in rows[:3] for v in row if v
            ).lower()
            scan_text += " " + sheet_name.lower()
            if "air" in scan_text:
                sheet_key = "air"
            elif "sea" in scan_text:
                sheet_key = "sea"
            else:
                print(f"   ⚠️  Unrecognized sheet '{sheet_name}' — skipping (no air/sea indicator found)")
                continue

            # Find header row (the one containing "Code")
            header_idx = None
            for i, row in enumerate(rows):
                if row and any(str(v).strip() == "Code" for v in row if v):
                    header_idx = i
                    break
            if header_idx is None:
                continue

            header = rows[header_idx]
            col = {str(h).strip(): i for i, h in enumerate(header) if h}
            code_i = col.get("Code")
            desc_i = col.get("Description")
            if code_i is None:
                continue

            alloc_idxs = [col[c] for c in STORE_AND_WAREHOUSE_COLS if c in col]
            if not alloc_idxs:
                alloc_idxs = None
                qty_i = col.get("Quantity")

            for row in rows[header_idx + 1:]:
                if not row or all(v is None for v in row):
                    continue
                code = row[code_i]
                if not code:
                    continue
                code = str(code).strip().upper()

                if alloc_idxs:
                    qty = sum(
                        int(row[i]) for i in alloc_idxs
                        if row[i] is not None and str(row[i]).strip() != ""
                    )
                else:
                    raw_qty = row[qty_i] if qty_i is not None else None
                    if not raw_qty:
                        continue
                    try:
                        qty = int(raw_qty)
                    except (ValueError, TypeError):
                        continue

                if qty > 0:
                    qty_by_sku[code][sheet_key] += qty
                    if desc_i is not None and row[desc_i] and not qty_by_sku[code]["description"]:
                        qty_by_sku[code]["description"] = str(row[desc_i]).strip()
    except Exception as e:
        print(f"   ⚠️  Excel parse error: {e}")

    # Flatten into plain dicts with a computed total
    result = {}
    for sku, d in qty_by_sku.items():
        result[sku] = {
            "air": d["air"], "sea": d["sea"], "total": d["air"] + d["sea"],
            "description": d["description"],
        }
    return result


# ─── FETCH SENT ORDER EMAILS ───────────────────────────────────────────────────
def _normalize_subject(subj):
    """
    Strip Re:/Fwd:/Fw: prefixes (possibly repeated, e.g. "Re: Re: ...") and
    surrounding whitespace so replies to the same order thread collapse to
    one key. This prevents the same order spreadsheet — re-sent as a reply —
    from being counted twice.
    """
    s = subj.strip()
    prefix_re = re.compile(r"^(re|fwd?|aw)\s*:\s*", re.IGNORECASE)
    while True:
        new_s = prefix_re.sub("", s).strip()
        if new_s == s:
            break
        s = new_s
    return s.lower()


def fetch_order_emails(token, days_back):
    """
    Search the Sent Items folder for weekly order emails matching SUBJECT_HINTS,
    within the last `days_back` days. Returns list of:
      {sent_date: datetime, subject: str, items: {sku: qty}}

    Replies to the same order thread (e.g. "Re: IKON WK23&24 Weekly order")
    carry the same attachment as the original and must NOT be treated as a
    second, separate order — only the latest email per normalized-subject
    thread is parsed.
    """
    since = (datetime.now() - timedelta(days=days_back)).strftime("%Y-%m-%dT00:00:00Z")
    filter_q  = urllib.parse.quote(f"sentDateTime ge {since}")
    orderby_q = urllib.parse.quote("sentDateTime desc")
    url = (f"{GRAPH_BASE}/me/mailFolders/SentItems/messages"
           f"?$filter={filter_q}"
           f"&$select=id,subject,sentDateTime,hasAttachments"
           f"&$orderby={orderby_q}"
           f"&$top=50")

    data = graph_get(token, url)
    candidates = []
    for m in data.get("value", []):
        if not m.get("hasAttachments"):
            continue
        subj = (m.get("subject") or "").lower()
        if any(hint in subj for hint in SUBJECT_HINTS):
            candidates.append(m)

    print(f"   Found {len(candidates)} order emails (incl. replies) in the last {days_back} days")

    # Collapse reply threads: keep only the most recent email per normalized subject.
    # candidates are already sorted sentDateTime desc, so the first one seen per
    # thread key is the latest.
    latest_per_thread = {}
    for m in candidates:
        key = _normalize_subject(m.get("subject", ""))
        if key not in latest_per_thread:
            latest_per_thread[key] = m
    deduped = list(latest_per_thread.values())

    if len(deduped) < len(candidates):
        print(f"   Collapsed to {len(deduped)} unique order thread(s) "
              f"(ignored {len(candidates)-len(deduped)} reply/duplicate email(s))")

    orders = []
    for msg in deduped:
        atts = graph_get(token, f"{GRAPH_BASE}/me/messages/{msg['id']}/attachments").get("value", [])
        for att in atts:
            name = att.get("name", "")
            if not name.lower().endswith((".xlsx", ".xls")):
                continue
            excel_bytes = base64.b64decode(att.get("contentBytes", ""))
            items = parse_order_excel(excel_bytes)
            if not items:
                continue
            sent_dt = datetime.strptime(msg["sentDateTime"][:19], "%Y-%m-%dT%H:%M:%S")
            orders.append({
                "sent_date": sent_dt,
                "subject":   msg.get("subject", ""),
                "items":     items,
            })
            print(f"   📧 {msg.get('subject','')[:50]:<50} | {sent_dt.strftime('%Y-%m-%d')} | {len(items)} SKUs")
            break  # one order workbook per email is enough

    return orders


# ─── BUILD INVOICE QUANTITY TIMELINE ───────────────────────────────────────────
def build_invoice_timeline(db):
    """
    Returns {sku: [(invoice_date, qty), ...]} from all shipments in shipment_db.json.
    invoice_date is parsed from the shipment's 'created' or 'eta' field as a proxy
    for "when this invoice was received" — falls back to None if unparseable.
    """
    timeline = defaultdict(list)
    shipments = db.get("shipments", {})

    for inv_no, shp in shipments.items():
        # Use whichever date field looks like an invoice/creation date
        date_str = shp.get("invoice_date") or shp.get("created") or shp.get("eta") or ""
        inv_date = None
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
            try:
                inv_date = datetime.strptime(date_str[:19], fmt)
                break
            except (ValueError, TypeError):
                continue

        for item in shp.get("items", []):
            sku = str(item.get("sku", "")).strip().upper()
            qty = item.get("quantity", 0) or 0
            if sku and qty:
                timeline[sku].append((inv_date, qty))

    return timeline


# ─── MATCH ORDERS TO INVOICES (GLOBAL FIFO PER SKU) ────────────────────────────
def compute_pending(orders, invoice_timeline):
    """
    FIFO matching per SKU across ALL orders, not per-order-independent matching.

    Problem with naive approach: if SKU X is ordered twice (WK23 and WK24), and
    one invoice arrives after WK22, comparing each order independently against
    "all invoices after THIS order's date" lets the SAME invoice quantity get
    claimed by both orders. That double-counts fulfillment.

    Fix: for each SKU, collect ALL orders (sorted by date) and ALL invoice lines
    (sorted by date) globally. Walk through orders oldest-first; each order
    consumes invoice quantity from a single shared pool, but only from invoices
    dated on/after THAT order's send date. Once an invoice unit is consumed by
    an earlier order, it can't be reused by a later order for the same SKU.
    """
    now = datetime.now()

    # Group all order-lines by SKU: [(order_date, qty, air, sea, subject), ...]
    orders_by_sku = defaultdict(list)
    for order in orders:
        for sku, qty_info in order["items"].items():
            orders_by_sku[sku].append({
                "order_date": order["sent_date"],
                "qty":        qty_info["total"],
                "air":        qty_info.get("air", 0),
                "sea":        qty_info.get("sea", 0),
                "description": qty_info.get("description", ""),
                "subject":    order["subject"],
            })

    pending = []

    for sku, sku_orders in orders_by_sku.items():
        # Oldest order first — earlier orders get first claim on later invoices
        sku_orders = sorted(sku_orders, key=lambda o: o["order_date"])

        # Invoice lines for this SKU, oldest first: [(date, qty), ...]
        invoice_lines = sorted(
            [(d, q) for (d, q) in invoice_timeline.get(sku, []) if d is not None],
            key=lambda x: x[0]
        )
        # Mutable remaining-quantity pool per invoice line (by index)
        invoice_remaining = [q for (_, q) in invoice_lines]

        for o in sku_orders:
            order_date = o["order_date"]
            need       = o["qty"]
            consumed   = 0

            # Consume from invoices dated on/after this order's send date,
            # oldest-eligible-invoice first, until `need` is satisfied or
            # invoices run out.
            for i, (inv_date, _) in enumerate(invoice_lines):
                if need <= 0:
                    break
                if inv_date < order_date:
                    continue  # invoice predates this order — can't count
                avail = invoice_remaining[i]
                if avail <= 0:
                    continue
                take = min(avail, need)
                invoice_remaining[i] -= take
                need     -= take
                consumed += take

            remaining = o["qty"] - consumed
            if remaining > 0:
                days_pending = (now - order_date).days
                # Proportionally split the still-outstanding qty by shipping method,
                # since invoices aren't tagged air/sea individually — we only know
                # the total invoiced amount, not which leg it came from.
                ordered_air, ordered_sea = o["air"], o["sea"]
                if o["qty"] > 0:
                    remaining_air = round(remaining * ordered_air / o["qty"])
                    remaining_sea = remaining - remaining_air
                else:
                    remaining_air = remaining_sea = 0

                pending.append({
                    "sku":               sku,
                    "order_description": o.get("description", ""),
                    "ordered_qty":       o["qty"],
                    "ordered_air":       ordered_air,
                    "ordered_sea":       ordered_sea,
                    "invoiced_qty":      consumed,
                    "remaining_qty":     remaining,
                    "remaining_air":     remaining_air,
                    "remaining_sea":     remaining_sea,
                    "order_date":        order_date.strftime("%Y-%m-%d"),
                    "order_subject":     o["subject"],
                    "days_pending":      days_pending,
                    "overdue":           days_pending > OVERDUE_DAYS,
                })

    pending.sort(key=lambda x: (-x["overdue"], -x["days_pending"]))
    return pending


# ─── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Track pending IKON orders not yet invoiced")
    parser.add_argument("--days-back", type=int, default=35,
                        help="How many days of sent order emails to check (default 35)")
    args = parser.parse_args()

    print("=" * 60)
    print("  IKON Pending Order Tracker")
    print("=" * 60)

    with open(TOKEN_FILE) as f:
        saved = json.load(f)
    token = do_refresh(saved)
    print("✅ Authenticated\n")

    print("📧 Searching Sent Items for weekly order emails...")
    orders = fetch_order_emails(token, args.days_back)
    if not orders:
        print("\n❌ No order emails with Excel attachments found.")
        return

    print(f"\n📊 Loading {DB_FILE} for invoice history...")
    with open(DB_FILE) as f:
        db = json.load(f)
    invoice_timeline = build_invoice_timeline(db)
    print(f"   Tracked {len(invoice_timeline)} unique SKUs across all invoices")

    print("\n🔍 Matching orders against invoices received after each order date...")
    pending = compute_pending(orders, invoice_timeline)

    overdue_count = sum(1 for p in pending if p["overdue"])
    print(f"\n📋 Result: {len(pending)} SKU(s) pending invoice ({overdue_count} overdue > {OVERDUE_DAYS} days)")

    for p in pending[:20]:
        flag = "🔴 OVERDUE" if p["overdue"] else "🟡 pending"
        ship_str = f"✈️{p['remaining_air']}/🚢{p['remaining_sea']}"
        print(f"   {flag} | {p['sku']:<12} | ordered {p['ordered_qty']:>4} | "
              f"invoiced {p['invoiced_qty']:>4} | remaining {p['remaining_qty']:>4} ({ship_str}) | "
              f"{p['days_pending']}d ago | {p['order_date']}")

    # Save output
    output = {
        "generated_at": datetime.now().isoformat(),
        "overdue_days_threshold": OVERDUE_DAYS,
        "pending": pending,
    }
    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, indent=2)
    print(f"\n✅ Saved → {OUTPUT_FILE}")

    # Copy to Streamlit data folder
    if os.path.exists(STREAMLIT_DATA_DIR):
        import shutil
        dest = os.path.join(STREAMLIT_DATA_DIR, OUTPUT_FILE)
        shutil.copy(OUTPUT_FILE, dest)
        print(f"✅ Copied to {dest}")
        print("   → Commit & Push to update the live dashboard!")


if __name__ == "__main__":
    main()
