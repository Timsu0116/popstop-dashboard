import streamlit as st
import pandas as pd
import os
import pickle
import json
from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR    = os.path.join(os.path.dirname(__file__), "data")
LIMITS_FILE = os.path.join(os.path.dirname(__file__), "store_limits.csv")

st.set_page_config(
    page_title="PopStop Retail Dashboard",
    page_icon="🛍️",
    layout="wide"
)

XL_COLORS = {
    "l1w_bg":      "E3F2FD",
    "l3w_bg":      "E8F5E9",
    "l6w_bg":      "FFF3E0",
    "woh_bg":      "F3E5F5",
    "higher_good": "C8E6C9",
    "lower_bad":   "FFCDD2",
    "header_bg":   "37474F",
    "header_font": "FFFFFF",
    "row_alt":     "F5F5F5",
    "border":      "BDBDBD",
    "yellow":      "FFF9C4",
}

USERS = {
    "admin":        {"password": "popstop2026",  "role": "admin",   "store": None},
    "manager":      {"password": "manager2026",  "role": "manager", "store": None},
    "dunedin":      {"password": "dun2026",      "role": "store",   "store": "Dunedin"},
    "papanui":      {"password": "pap2026",      "role": "store",   "store": "Papanui"},
    "riccarton":    {"password": "ricc2026",     "role": "store",   "store": "Riccarton"},
    "queensgate":   {"password": "qg2026",       "role": "store",   "store": "Queensgate"},
    "richmond":     {"password": "rich2026",     "role": "store",   "store": "Richmond"},
    "sylviapark":   {"password": "sp2026",       "role": "store",   "store": "SylviaPark"},
    "terapa":       {"password": "tr2026",       "role": "store",   "store": "Terapa"},
    "office":       {"password": "office2026",   "role": "store",   "store": "Office"},
    "warehouseakl": {"password": "akl2026",      "role": "store",   "store": "WarehouseAKL"},
}

# ============================================================
# 加载数据
# ============================================================
def _get_mtime(filepath):
    """返回文件修改时间戳，文件不存在则返回0。用作cache key。"""
    try:
        return os.path.getmtime(filepath)
    except:
        return 0

@st.cache_data
def load_company(_mtime=0):
    df = pd.read_parquet(os.path.join(DATA_DIR, "output_company.parquet"))
    date_cols = ["created","first_sale","last_sale","last_received"]
    for c in date_cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce").dt.strftime("%Y-%m-%d")
    for c in df.select_dtypes(include="number").columns:
        if c in ["supply_price","retail_price","avg_weekly","WOH",
                 "inventory_cost","retail_value"]:
            df[c] = df[c].round(2)
        elif c in ["days_since_created","days_since_first_sale",
                   "days_since_last_sold","days_since_last_received"]:
            df[c] = df[c].round(0).astype("Int64")
        elif c in ["DUN","PAP","QG","RICC","RICH","SP","TR","Office","WH1","AKLWH","TOT",
                   "DUN_L1W","PAP_L1W","QG_L1W","RICC_L1W","RICH_L1W","SP_L1W",
                   "TR_L1W","Office_L1W","STORE_TOT","items_sold_lifetime","CY26",
                   "L6W","L3W","L1W"]:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)
        else:
            df[c] = df[c].round(1)
    return df

@st.cache_data
def load_store(_mtime=0):
    df = pd.read_parquet(os.path.join(DATA_DIR, "output_store.parquet"))
    date_cols = ["created","first_sale","last_sale","last_received"]
    for c in date_cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce").dt.strftime("%Y-%m-%d")
    for c in df.select_dtypes(include="number").columns:
        if c in ["supply_price","retail_price","avg_weekly","avg_weekly_co",
                 "WOH_store","WOH_co","inventory_cost","retail_value"]:
            df[c] = df[c].round(2)
        elif c in ["days_since_last_sold","days_since_last_received"]:
            df[c] = df[c].round(0).astype("Int64")
        else:
            df[c] = df[c].round(1)
    return df

@st.cache_data
def load_limits(_mtime=0):
    if os.path.exists(LIMITS_FILE):
        return pd.read_csv(LIMITS_FILE)
    return pd.DataFrame(columns=["store","category","max_sku","max_units"])

@st.cache_data
def load_preorders(_mtime=0):
    path = os.path.join(DATA_DIR, "preorders.json")
    if not os.path.exists(path):
        return {}
    with open(path) as f:
        return json.load(f)

@st.cache_data
def load_pending_orders_map(_mtime=0):
    """
    {supplier_code: {"remaining_qty": int, "remaining_air": int, "remaining_sea": int,
                      "days_pending": int, "overdue": bool}}
    Aggregated across multiple pending order lines for the same SKU (sums quantities,
    keeps the longest days_pending / any overdue flag).
    """
    path = os.path.join(DATA_DIR, "pending_orders.json")
    if not os.path.exists(path):
        return {}
    with open(path) as f:
        data = json.load(f)
    out = {}
    for p in data.get("pending", []):
        sku = str(p.get("sku", "")).strip().upper()
        if not sku:
            continue
        if sku not in out:
            out[sku] = {"remaining_qty": 0, "remaining_air": 0, "remaining_sea": 0,
                       "days_pending": 0, "overdue": False}
        out[sku]["remaining_qty"] += p.get("remaining_qty", 0)
        out[sku]["remaining_air"] += p.get("remaining_air", 0)
        out[sku]["remaining_sea"] += p.get("remaining_sea", 0)
        out[sku]["days_pending"]   = max(out[sku]["days_pending"], p.get("days_pending", 0))
        out[sku]["overdue"]        = out[sku]["overdue"] or p.get("overdue", False)
    return out

@st.cache_data
def load_shipment_db(_mtime=0):
    db_path = os.path.join(DATA_DIR, "shipment_db.json")
    if not os.path.exists(db_path):
        return {}
    with open(db_path) as f:
        db = json.load(f)
    return db.get("shipments", {})

def build_transit_map(shipments):
    """Build {supplier_code: 'QG:12 RICC:6 ...'} from shipment_db"""
    from collections import defaultdict
    # store_abbrev map
    STORE_ABBREV = {
        "Dunedin":       "DUN",
        "Papanui":       "PAP",
        "Queensgate":    "QG",
        "Riccarton":     "RICC",
        "Richmond":      "RICH",
        "Sylvia Park":   "SP",
        "Te Rapa":       "TR",
        "Office":        "OFC",
        "Warehouse AKL": "AKL",
        "WH1":           "WH1",
    }
    ACTIVE = {"Pending","Waiting for Shipping","In Transit","Documents Received","In Customs","Arrived NZ","Customs Cleared"}
    # sku -> store -> qty
    transit = defaultdict(lambda: defaultdict(int))
    for inv_no, shp in shipments.items():
        status = shp.get("tnl_status", "Pending")
        if status not in ACTIVE:
            continue
        store = shp.get("store", "")
        abbrev = STORE_ABBREV.get(store, store[:4])
        for item in shp.get("items", []):
            sku = item.get("sku", "").strip().upper()
            qty = item.get("quantity", 0) or 0
            if sku and qty:
                transit[sku][abbrev] += qty
    # Format: "QG:24 RICC:12"
    result = {}
    for sku, stores in transit.items():
        parts = [f"{st}:{qty}" for st, qty in sorted(stores.items())]
        result[sku] = " ".join(parts)
    return result

@st.cache_data
def load_weekly(_mtime=0):
    pkl_file = os.path.join(DATA_DIR, "weekly_reports.pkl")
    if os.path.exists(pkl_file):
        with open(pkl_file, "rb") as f:
            return pickle.load(f)
    return None

# 预加载（传入文件修改时间，文件更新后缓存自动失效）
with st.spinner("🔄 Initializing..."):
    _co_mtime  = _get_mtime(os.path.join(DATA_DIR, "output_company.parquet"))
    _st_mtime  = _get_mtime(os.path.join(DATA_DIR, "output_store.parquet"))
    _lim_mtime = _get_mtime(LIMITS_FILE)
    _pkl_mtime = _get_mtime(os.path.join(DATA_DIR, "weekly_reports.pkl"))
    _po_mtime  = _get_mtime(os.path.join(DATA_DIR, "preorders.json"))
    load_company(_mtime=_co_mtime)
    load_store(_mtime=_st_mtime)
    load_limits(_mtime=_lim_mtime)
    load_weekly(_mtime=_pkl_mtime)
    load_preorders(_mtime=_po_mtime)

# ============================================================
# 登录
# ============================================================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.user  = None
    st.session_state.role  = None
    st.session_state.store = None

if not st.session_state.logged_in:
    st.title("🛍️ PopStop Retail Dashboard")
    st.markdown("---")
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.subheader("Login")
        username = st.text_input("Username", key="login_user").lower()
        password = st.text_input("Password", type="password", key="login_pass")
        if st.button("Login", use_container_width=True, key="login_btn"):
            if username in USERS and USERS[username]["password"] == password:
                st.session_state.logged_in = True
                st.session_state.user  = username
                st.session_state.role  = USERS[username]["role"]
                st.session_state.store = USERS[username]["store"]
                st.rerun()
            else:
                st.error("Invalid username or password")
    st.stop()

# ============================================================
# Excel工具
# ============================================================
def _get_border():
    side = Side(style="thin", color=XL_COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)

def _format_sheet(ws, df, is_store=False):
    """格式化Excel - 按样本文件格式"""
    from openpyxl.styles import Alignment as XLAlignment
    col_names = list(df.columns)
    border    = _get_border()
    center_align = XLAlignment(horizontal="center", vertical="center")
    left_align   = XLAlignment(horizontal="left",   vertical="center")

    def get_ci(name):
        try: return col_names.index(name) + 1
        except: return None

    GROUP_COLORS = {
        "id":"37474F","product":"37474F","sku":"37474F","category":"37474F",
        "tag":"37474F","supply_price":"37474F","retail_price":"37474F",
        "inventory_cost":"37474F","retail_value":"37474F","brand":"37474F",
        "supplier":"37474F","supplier_code":"37474F","store":"37474F",
        "created":"37474F","first_sale":"37474F","last_sale":"37474F",
        "last_received":"37474F","days_since_created":"37474F",
        "days_since_first_sale":"37474F","days_since_last_sold":"37474F",
        "days_since_last_received":"37474F",
        "DUN":"1565C0","PAP":"1565C0","QG":"1565C0","RICC":"1565C0",
        "RICH":"1565C0","SP":"1565C0","TR":"1565C0","Office":"1565C0",
        "WH1":"1565C0","AKLWH":"1565C0","TOT":"1565C0",
        "soh":"1565C0","SOH_co":"1565C0",
        "Pre-order":"6A1B9A","In Transit":"6A1B9A",
        "DUN_L1W":"2E7D32","PAP_L1W":"2E7D32","QG_L1W":"2E7D32",
        "RICC_L1W":"2E7D32","RICH_L1W":"2E7D32","SP_L1W":"2E7D32",
        "TR_L1W":"2E7D32","Office_L1W":"2E7D32","STORE_TOT":"2E7D32",
        "items_sold_lifetime":"4E342E","CY26":"4E342E","items_sold":"4E342E",
        "items_sold_lifetime_co":"4E342E",
        "avg_weekly":"4E342E","avg_weekly_co":"4E342E",
        "WOH":"4E342E","WOH_store":"4E342E","WOH_co":"4E342E",
        "L6W":"4E342E","L6W_co":"4E342E","L3W":"4E342E","L3W_co":"4E342E",
        "L1W":"4E342E","L1W_co":"4E342E",
        "Company_Status":"4A148C","Action":"4A148C","Store_Status":"4A148C",
    }
    HIDDEN_CO = {"id","tag","supply_price","inventory_cost","supplier_code",
                 "created","days_since_created"}
    HIDDEN_ST = {"store","tag","supply_price","inventory_cost","supplier_code",
                 "created","first_sale","CY26","items_sold",
                 "items_sold_lifetime_co","days_since_created"}
    hidden = HIDDEN_ST if is_store else HIDDEN_CO
    ZERO_BLANK = {
        "DUN","PAP","QG","RICC","RICH","SP","TR","Office","WH1","AKLWH",
        "DUN_L1W","PAP_L1W","QG_L1W","RICC_L1W","RICH_L1W",
        "SP_L1W","TR_L1W","Office_L1W","STORE_TOT",
        "Pre-order","soh","SOH_co",
        "L1W","L1W_co","L3W","L3W_co","L6W","L6W_co",
        "items_sold","items_sold_lifetime","items_sold_lifetime_co",
    }
    COL_WIDTHS = {
        "product":36,"sku":9,"category":13,"brand":10,"supplier":15,
        "retail_price":9,"retail_value":12,
        "DUN":8,"PAP":8,"QG":8,"RICC":8,"RICH":8,"SP":8,"TR":8,
        "Office":8,"WH1":8,"AKLWH":8,"TOT":9,
        "Pre-order":11,"In Transit":25,
        "last_sale":11,"last_received":13,"first_sale":11,
        "days_since_last_sold":10,"days_since_last_received":10,
        "days_since_first_sale":10,
        "DUN_L1W":8,"PAP_L1W":8,"QG_L1W":8,"RICC_L1W":8,"RICH_L1W":8,
        "SP_L1W":8,"TR_L1W":8,"Office_L1W":8,"STORE_TOT":9,
        "items_sold_lifetime":10,"CY26":8,
        "avg_weekly":9,"avg_weekly_co":9,
        "WOH":7,"WOH_store":8,"WOH_co":8,
        "L6W":7,"L6W_co":7,"L3W":7,"L3W_co":7,"L1W":7,"L1W_co":7,
        "Company_Status":13,"Action":16,"Store_Status":13,
        "soh":8,"SOH_co":8,"store":12,"supplier_code":14,
    }
    TEXT_COLS = {"product","tag","brand","supplier","In Transit","store"}
    ACTION_COLORS = {
        "URGENT RESTOCK": "EF9A9A",
        "RESTOCK":        "C8E6C9",
        "REPLENISH":      "A5D6A7",
        "TRANSFER OUT":   "FFE082",
        "REDUCE / CLR":   "FFE0B2",
        "CLEAR":          "FFCDD2",
        "REDUCE ORDER":   "F8BBD9",
        "CHECK":          "B2EBF2",
        "MONITOR":        "F5F5F5",
    }

    ws.row_dimensions[1].height = 34.5
    for i, col_name in enumerate(col_names, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width  = COL_WIDTHS.get(col_name, 9)
        ws.column_dimensions[letter].hidden = col_name in hidden
        hdr_color = GROUP_COLORS.get(col_name, "37474F")
        if col_name.startswith("2026-W") or col_name.startswith("2025-W"):
            hdr_color = "546E7A"
        cell = ws.cell(row=1, column=i)
        cell.value     = col_name
        cell.fill      = PatternFill("solid", fgColor=hdr_color)
        cell.font      = Font(color="FFFFFF", bold=True, size=9)
        cell.border    = border
        cell.alignment = center_align

    # Both company and store sheets: only row height, no per-cell ops to avoid OOM
    # Store sheets can have 500-800 rows × many columns = hundreds of thousands of cell ops
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 18

    ai = get_ci("Action")
    if ai:
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=ai).value
            if val in ACTION_COLORS:
                ws.cell(row=row, column=ai).fill = PatternFill("solid", fgColor=ACTION_COLORS[val])
                ws.cell(row=row, column=ai).font = Font(bold=True, size=9)

    ws.freeze_panes = "C2"


def generate_excel(co_df, store_df, limits_df, store_name=None):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Reorder: move Pre-order, In Transit, Pending Invoice to after TOT column
        co_export = co_df.copy()
        if "TOT" in co_export.columns:
            tot_idx = co_export.columns.tolist().index("TOT")
            extra_cols = [c for c in ["Pre-order","In Transit","Pending Invoice"] if c in co_export.columns]
            other_cols  = [c for c in co_export.columns if c not in extra_cols]
            insert_at   = other_cols.index("TOT") + 1
            new_order   = other_cols[:insert_at] + extra_cols + other_cols[insert_at:]
            co_export   = co_export[new_order]
        co_export.to_excel(writer, index=False, sheet_name="Company")
        _format_sheet(writer.sheets["Company"], co_export, is_store=False)

        if store_name and store_name not in ["All", "skip", None]:
            sdf = store_df[store_df["store"] == store_name] if "store" in store_df.columns else store_df
            if not sdf.empty:
                sdf.to_excel(writer, index=False, sheet_name=store_name[:28])
                _format_sheet(writer.sheets[store_name[:28]], sdf, is_store=True)
            if not limits_df.empty:
                ldf = limits_df[limits_df["store"] == store_name]
                if not ldf.empty:
                    ldf.to_excel(writer, index=False, sheet_name=f"{store_name[:20]}_Limits")

    return buffer.getvalue()


def write_report_sheet(writer, sheet_name, report, weeks, include_sku=False, include_pct=False):
    metrics = [
        ("REV $",    "REV",    False, "#,##0.00"),
        ("GP $",     "GP",     False, "#,##0.00"),
        ("MARGIN %", "MARGIN", True,  "0.00%"),
        ("SALES U",  "SALES_U",False, "#,##0"),
    ]
    if include_sku:
        metrics.append(("SKU Count","SKU",False,"#,##0"))
    if include_pct:
        metrics.append(("REV % of Category","REV_PCT",True,"0.00%"))

    wb     = writer.book
    ws     = wb.create_sheet(title=sheet_name[:31])
    border = _get_border()

    header_fill = PatternFill("solid", fgColor=XL_COLORS["header_bg"])
    header_font = Font(color=XL_COLORS["header_font"], bold=True)
    section_fills = {
        "REV $":              PatternFill("solid", fgColor="E3F2FD"),
        "GP $":               PatternFill("solid", fgColor="E8F5E9"),
        "MARGIN %":           PatternFill("solid", fgColor="FFF3E0"),
        "SALES U":            PatternFill("solid", fgColor="F3E5F5"),
        "SKU Count":          PatternFill("solid", fgColor="FCE4EC"),
        "REV % of Category":  PatternFill("solid", fgColor="E8EAF6"),
    }

    current_row = 1
    for metric_label, metric_key, is_pct, num_fmt in metrics:
        if metric_key not in report:
            continue
        df        = report[metric_key].copy()
        group_col = df.columns[0]

        ws.cell(row=current_row, column=1, value=metric_label)
        ws.cell(row=current_row, column=1).fill = section_fills.get(metric_label, PatternFill())
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1

        ws.cell(row=current_row, column=1, value="Brand / Tag / Category")
        ws.cell(row=current_row, column=1).fill   = header_fill
        ws.cell(row=current_row, column=1).font   = header_font
        ws.cell(row=current_row, column=1).border = border

        for ci, wk in enumerate(weeks, start=2):
            cell = ws.cell(row=current_row, column=ci, value=wk)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.border = border

        tot_ci = len(weeks) + 2
        if not is_pct:
            cell = ws.cell(row=current_row, column=tot_ci, value="TOTAL")
            cell.fill   = header_fill
            cell.font   = header_font
            cell.border = border
        current_row += 1

        alt_fill  = PatternFill("solid", fgColor=XL_COLORS["row_alt"])
        norm_fill = PatternFill("solid", fgColor="FFFFFF")

        for ri, data_row in df.iterrows():
            fill = alt_fill if ri % 2 == 0 else norm_fill
            cell = ws.cell(row=current_row, column=1, value=data_row[group_col])
            cell.fill   = fill
            cell.border = border

            row_total = 0
            for ci, wk in enumerate(weeks, start=2):
                val  = data_row.get(wk, 0) or 0
                cell = ws.cell(row=current_row, column=ci,
                               value=round(val,4) if is_pct else round(val,2))
                cell.fill         = fill
                cell.border       = border
                cell.number_format = num_fmt
                if not is_pct:
                    row_total += val

            if not is_pct:
                cell = ws.cell(row=current_row, column=tot_ci, value=round(row_total,2))
                cell.fill          = fill
                cell.border        = border
                cell.font          = Font(bold=True)
                cell.number_format = num_fmt
            current_row += 1

        current_row += 1

    ws.column_dimensions["A"].width = 40
    for ci in range(2, len(weeks) + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    ws.freeze_panes = "B3"


def _safe_val(df, cat, grp_col, wk):
    try:
        row = df[df[grp_col] == cat]
        if row.empty or wk not in df.columns:
            return 0.0
        return float(row[wk].values[0]) or 0.0
    except:
        return 0.0


def write_category_summary_sheet(wb, cat_company, cat_stores, weeks):
    """Category Summary with REV/GP/Margin % - company and each store"""
    from openpyxl.styles import Alignment as XLAlign
    if "REV" not in cat_company or "GP" not in cat_company:
        return

    ws     = wb.create_sheet(title="Category Summary", index=0)
    border = _get_border()
    center = XLAlign(horizontal="center", vertical="center")
    left   = XLAlign(horizontal="left",   vertical="center")
    wks_display = sorted(weeks, reverse=True)[:8]

    rev_df  = cat_company["REV"].copy()
    gp_df   = cat_company["GP"].copy()
    grp_col = rev_df.columns[0]
    all_wks = [c for c in rev_df.columns if str(c).startswith("2026-W")]

    fixed_hdrs = ["Category","Total REV","REV %","Total GP","GP %","Margin"]
    wk_hdrs = []
    for wk in wks_display:
        wk_hdrs += [f"{wk} REV", f"{wk} GP", f"{wk} Mgn"]
    all_hdrs = fixed_hdrs + wk_hdrs
    n_cols   = len(all_hdrs)

    ws.row_dimensions[1].height = 28
    t = ws.cell(1, 1, "Category Performance Summary — Revenue, GP & Margin %")
    t.font = Font(bold=True, size=13, color="1565C0")
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(n_cols,30))

    VALID_STORES = ["Dunedin","Papanui","Queensgate","Riccarton",
                    "Richmond","Sylvia Park","Te Rapa","Office"]
    STORE_COLORS = {
        "Dunedin":"37474F","Papanui":"2E7D32","Queensgate":"1B5E20",
        "Riccarton":"880E4F","Richmond":"4A148C","Sylvia Park":"E65100",
        "Te Rapa":"0D47A1","Office":"546E7A"
    }
    SECTION_CONFIGS = [("COMPANY (PopStop)", "1565C0", cat_company)]
    for s in VALID_STORES:
        if s in cat_stores and "REV" in cat_stores[s]:
            SECTION_CONFIGS.append((s, STORE_COLORS.get(s,"546E7A"), cat_stores[s]))

    current_row = 3
    for section_title, color, report in SECTION_CONFIGS:
        s_rev_df  = report["REV"].copy()
        s_gp_df   = report.get("GP", pd.DataFrame())
        s_grp_col = s_rev_df.columns[0]
        s_all_wks = [c for c in s_rev_df.columns if str(c).startswith("2026-W")]
        s_rev_all = float(s_rev_df[s_all_wks].sum().sum()) if s_all_wks else 0
        s_gp_all  = float(s_gp_df[[c for c in s_gp_df.columns if str(c).startswith("2026-W")]].sum().sum()) if not s_gp_df.empty and s_all_wks else 0
        s_cats    = s_rev_df[s_grp_col].tolist()

        # Section header
        ws.row_dimensions[current_row].height = 22
        sc = ws.cell(current_row, 1, section_title)
        sc.font = Font(bold=True, size=11, color="FFFFFF")
        sc.fill = PatternFill("solid", fgColor=color)
        if n_cols > 1:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=min(n_cols,30))
        current_row += 1

        # Column headers
        ws.row_dimensions[current_row].height = 20
        for ci, h in enumerate(all_hdrs, 1):
            c = ws.cell(current_row, ci, h)
            c.fill = PatternFill("solid", fgColor=color)
            c.font = Font(color="FFFFFF", bold=True, size=9)
            c.border = border
            c.alignment = left if ci == 1 else center
        current_row += 1

        # Sort cats by total rev descending
        cat_totals = []
        for cat in s_cats:
            cat_rev = sum(_safe_val(s_rev_df, cat, s_grp_col, wk) for wk in s_all_wks)
            cat_gp  = sum(_safe_val(s_gp_df, cat, s_gp_df.columns[0] if not s_gp_df.empty else s_grp_col, wk) for wk in s_all_wks)
            cat_totals.append((cat_rev, cat, cat_gp))
        cat_totals.sort(reverse=True)

        for ri, (cat_rev_tot, cat, cat_gp_tot) in enumerate(cat_totals):
            cat_mg  = cat_gp_tot / cat_rev_tot if cat_rev_tot else 0
            rev_pct = cat_rev_tot / s_rev_all if s_rev_all else 0
            gp_pct  = cat_gp_tot  / s_gp_all  if s_gp_all  else 0
            fill    = PatternFill("solid", fgColor="F8F9FA" if ri % 2 == 0 else "FFFFFF")
            ws.row_dimensions[current_row].height = 18

            vals = [cat, cat_rev_tot, rev_pct, cat_gp_tot, gp_pct, cat_mg]
            fmts = [None, "#,##0", "0.0%", "#,##0", "0.0%", "0.0%"]
            for ci, (v, f) in enumerate(zip(vals, fmts), 1):
                c = ws.cell(current_row, ci, v if v != 0 else (v if ci == 1 else None))
                c.fill = fill; c.border = border
                c.alignment = left if ci == 1 else center
                c.font = Font(size=9, bold=(ci in (3,5) and isinstance(v,float) and v >= 0.15),
                              color=("1B5E20" if ci in (3,5) and isinstance(v,float) and v >= 0.15 else "000000"))
                if f and isinstance(v, float): c.number_format = f

            ci = 7
            for wk in wks_display:
                wk_rev = _safe_val(s_rev_df, cat, s_grp_col, wk)
                s_gp_grp = s_gp_df.columns[0] if not s_gp_df.empty else s_grp_col
                wk_gp  = _safe_val(s_gp_df,  cat, s_gp_grp, wk)
                wk_mg  = wk_gp / wk_rev if wk_rev else 0
                for v, f in [(wk_rev,"#,##0"),(wk_gp,"#,##0"),(wk_mg,"0.0%")]:
                    c = ws.cell(current_row, ci, v if v else None)
                    c.fill = fill; c.border = border
                    c.alignment = center; c.font = Font(size=9)
                    if f and v: c.number_format = f
                    ci += 1
            current_row += 1

        # Totals
        tf = PatternFill("solid", fgColor=color)
        for ci, (v, f) in enumerate(zip(
            ["TOTAL", s_rev_all, 1.0, s_gp_all, 1.0, s_gp_all/s_rev_all if s_rev_all else 0],
            [None, "#,##0", "0.0%", "#,##0", "0.0%", "0.0%"]), 1):
            c = ws.cell(current_row, ci, v)
            c.fill = tf; c.font = Font(color="FFFFFF", bold=True, size=9)
            c.border = border; c.alignment = left if ci == 1 else center
            if f and isinstance(v, float): c.number_format = f
        ws.row_dimensions[current_row].height = 20
        current_row += 2

    ws.column_dimensions["A"].width = 28
    for ci, w in enumerate([12,9,12,9,9], 2):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ci in range(7, n_cols+1):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.freeze_panes = "B4"


def write_trend_chart_sheet(wb, cat_company, weeks):
    """Revenue trend line chart - placed at A2 next to data"""
    from openpyxl.chart import LineChart, Reference, Series

    if "REV" not in cat_company:
        return

    ws      = wb.create_sheet(title="Revenue Trends", index=1)
    rev_df  = cat_company["REV"].copy()
    grp_col = rev_df.columns[0]
    wks     = sorted([w for w in weeks])
    n_wks   = len(wks)

    # Write data table
    ws.cell(1, 1, "Category")
    for ci, wk in enumerate(wks, 2):
        ws.cell(1, ci, wk)
        ws.column_dimensions[get_column_letter(ci)].width = 11

    categories = rev_df[grp_col].tolist()
    cat_totals = []
    for ri, cat in enumerate(categories, 2):
        ws.cell(ri, 1, cat)
        total = 0
        cat_row = rev_df[rev_df[grp_col] == cat]
        for ci, wk in enumerate(wks, 2):
            val = float(cat_row[wk].values[0]) if wk in cat_row.columns and not cat_row.empty else 0
            ws.cell(ri, ci, round(val, 2))
            total += val
        cat_totals.append((total, ri, cat))

    # Top 15 by revenue
    cat_totals.sort(reverse=True)
    top_rows = [(r, c) for _, r, c in cat_totals[:15]]
    n_cats   = len(categories)

    # Build chart
    chart = LineChart()
    chart.title        = "Category Revenue Trend 2026 (Top 15)"
    chart.style        = 10
    chart.y_axis.title = "Revenue ($NZD)"
    chart.x_axis.title = "Week"
    chart.height       = 18
    chart.width        = 32
    chart.y_axis.numFmt = "#,##0"
    chart.legend.position = "r"

    cats_ref = Reference(ws, min_col=2, max_col=n_wks+1, min_row=1, max_row=1)
    for ri, cat_name in top_rows:
        data_ref = Reference(ws, min_col=2, max_col=n_wks+1, min_row=ri, max_row=ri)
        series   = Series(data_ref, title=cat_name)
        series.smooth = True
        chart.series.append(series)
    chart.set_categories(cats_ref)

    # Place chart right after data table (row 2, col n_wks+3)
    chart_col = n_wks + 3
    ws.add_chart(chart, f"{get_column_letter(chart_col)}2")

    ws.column_dimensions["A"].width = 28
    ws.freeze_panes = "B2"


def write_cat_store_comparison_sheet(wb, cat_company, cat_stores, weeks):
    """新增：所有Category的Store Share和Co Share汇总，带条件格式热力图"""
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment as XLAlign

    if "REV" not in cat_company or not cat_stores:
        return

    VALID_STORES = ["Dunedin","Papanui","Queensgate","Riccarton",
                    "Richmond","Sylvia Park","Te Rapa","Office"]
    store_list = [s for s in VALID_STORES if s in cat_stores]

    rev_co  = cat_company["REV"].copy()
    grp_col = rev_co.columns[0]
    all_cats = sorted(rev_co[grp_col].dropna().tolist())
    wks_display = sorted(weeks, reverse=True)[:12]  # max 12 weeks in Excel

    def get_rev(report, cat, wk):
        if "REV" not in report: return 0.0
        df = report["REV"]; gc = df.columns[0]
        if wk not in df.columns: return 0.0
        row = df[df[gc] == cat]
        return float(row[wk].values[0]) if not row.empty else 0.0

    def get_store_total(report, wk):
        if "REV" not in report: return 0.0
        df = report["REV"]
        return float(df[wk].sum()) if wk in df.columns else 0.0

    def get_co_cat_rev(cat, wk):
        row = rev_co[rev_co[grp_col] == cat]
        if row.empty or wk not in rev_co.columns: return 0.0
        return float(row[wk].values[0])

    border  = _get_border()
    center  = XLAlign(horizontal="center", vertical="center")
    left    = XLAlign(horizontal="left",   vertical="center")

    # ── Sheet 1: Store Share % ────────────────────────────────────
    for sheet_idx, (metric_label, metric_suffix, clr_start, clr_mid, clr_end) in enumerate([
        ("Store Share %",   "st",  "F5F5F5", "81C784", "1B5E20"),
        ("Company Cat %",   "co",  "F5F5F5", "90CAF9", "1565C0"),
    ]):
        ws = wb.create_sheet(f"Cat_{metric_label.replace(' ','_')[:28]}")
        current_row = 1

        # Title
        t = ws.cell(1, 1, f"Category {metric_label} by Store — All Categories")
        t.font = Font(bold=True, size=12, color="1565C0")
        current_row = 3

        for cat in all_cats:
            # Section header
            ws.row_dimensions[current_row].height = 20
            sc = ws.cell(current_row, 1, cat)
            sc.font = Font(bold=True, size=10, color="FFFFFF")
            sc.fill = PatternFill("solid", fgColor="37474F")
            sc.border = border
            n_cols = len(wks_display) + 1
            if n_cols > 1:
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=n_cols)
            current_row += 1

            # Column headers
            ws.cell(current_row, 1, "Store").font = Font(bold=True, size=9)
            ws.cell(current_row, 1).fill   = PatternFill("solid", fgColor="ECEFF1")
            ws.cell(current_row, 1).border = border
            for ci, wk in enumerate(wks_display, 2):
                c = ws.cell(current_row, ci, wk.replace("2026-",""))
                c.font = Font(bold=True, size=9)
                c.fill = PatternFill("solid", fgColor="ECEFF1")
                c.border = border; c.alignment = center
            current_row += 1

            data_start = current_row
            for store in store_list:
                report = cat_stores.get(store, {})
                ws.cell(current_row, 1, store).border = border
                ws.cell(current_row, 1).font = Font(size=9)
                for ci, wk in enumerate(wks_display, 2):
                    rev        = get_rev(report, cat, wk)
                    if metric_suffix == "st":
                        total  = get_store_total(report, wk)
                        val    = rev / total if total > 0 else None
                    else:
                        co_rev = get_co_cat_rev(cat, wk)
                        val    = rev / co_rev if co_rev > 0 else None
                    c = ws.cell(current_row, ci, round(val, 4) if val else None)
                    c.number_format = "0.0%"; c.border = border
                    c.alignment = center; c.font = Font(size=9)
                current_row += 1

            # Color scale on data range
            last_col = get_column_letter(len(wks_display)+1)
            if current_row > data_start:
                ws.conditional_formatting.add(
                    f"B{data_start}:{last_col}{current_row-1}",
                    ColorScaleRule(
                        start_type="num",        start_value=0,  start_color=clr_start,
                        mid_type="percentile",   mid_value=50,   mid_color=clr_mid,
                        end_type="percentile",   end_value=100,  end_color=clr_end
                    )
                )
            current_row += 1  # blank row between categories

        ws.column_dimensions["A"].width = 18
        for ci in range(2, len(wks_display)+2):
            ws.column_dimensions[get_column_letter(ci)].width = 9
        ws.freeze_panes = "B4"


def generate_weekly_excel(weekly_data, weeks=None):
    if weeks is None:
        weeks = sorted(weekly_data["weeks"], reverse=True)
    cat_company   = weekly_data["cat_company"]
    cat_stores    = weekly_data["cat_stores"]
    sup_company   = weekly_data["sup_company"]
    brand_company = weekly_data.get("brand_company", {})
    tag_company   = weekly_data.get("tag_company", {})

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        write_category_summary_sheet(writer.book, cat_company, cat_stores, weeks)
        write_trend_chart_sheet(writer.book, cat_company, weeks)
        write_cat_store_comparison_sheet(writer.book, cat_company, cat_stores, weeks)
        write_report_sheet(writer, "Category_PopStop", cat_company, weeks)
        for s, report in cat_stores.items():
            write_report_sheet(writer, f"Cat_{s[:25]}", report, weeks)
        write_report_sheet(writer, "Supplier_PopStop", sup_company, weeks, include_sku=True)
        if brand_company:
            write_report_sheet(writer, "Brand_PopStop", brand_company, weeks, include_pct=True)
        if tag_company:
            write_report_sheet(writer, "Tag_PopStop", tag_company, weeks, include_pct=True)

    return buffer.getvalue()


def generate_category_detail_excel(weekly_data, category, weeks=None, scope="company", store_name=None):
    """生成某个Category的Brand/Tag细分报告"""
    if weeks is None:
        weeks = sorted(weekly_data["weeks"], reverse=True)
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if scope == "company":
            brand_data = weekly_data.get("brand_by_cat_company", {}).get(category)
            tag_data   = weekly_data.get("tag_by_cat_company", {}).get(category)
            if brand_data:
                write_report_sheet(writer, f"Brand_{category[:25]}", brand_data, weeks, include_pct=True)
            if tag_data:
                write_report_sheet(writer, f"Tag_{category[:25]}", tag_data, weeks, include_pct=True)
        else:
            brand_data = weekly_data.get("brand_by_cat_stores", {}).get(store_name, {}).get(category)
            tag_data   = weekly_data.get("tag_stores", {}).get(store_name)
            if brand_data:
                write_report_sheet(writer, f"Brand_{category[:25]}", brand_data, weeks, include_pct=True)
            if tag_data:
                write_report_sheet(writer, f"Tag_{store_name[:25]}", tag_data, weeks, include_pct=True)

    return buffer.getvalue()

# ============================================================
# Store Limits
# ============================================================
def calc_limits_comparison(store_df, limits_df, store_name):
    # Handle store name variants (e.g. "Sylvia Park" vs "SylviaPark")
    STORE_VARIANTS = {
        "Sylvia Park": ["Sylvia Park","SylviaPark","Sylvia_Park","sylviapark"],
        "Te Rapa":     ["Te Rapa","TeRapa","Te_Rapa","terapa"],
    }
    variants = STORE_VARIANTS.get(store_name, [store_name])
    store_data = store_df[store_df["store"].isin(variants)].copy()
    if store_data.empty:
        # Try case-insensitive match
        store_data = store_df[store_df["store"].str.lower() == store_name.lower()].copy()

    limits = limits_df[limits_df["store"] == store_name].copy()
    if limits.empty:
        return None
    actual = store_data.groupby("category").agg(
        actual_sku   = ("sku","nunique"),
        actual_units = ("soh","sum")
    ).reset_index()
    result = limits.merge(actual, on="category", how="left")
    result["actual_sku"]   = result["actual_sku"].fillna(0).astype(int)
    result["actual_units"] = result["actual_units"].fillna(0).astype(int)
    result["sku_diff"]     = result["max_sku"]   - result["actual_sku"]
    result["units_diff"]   = result["max_units"] - result["actual_units"]
    result["sku_pct"]   = result.apply(
        lambda r: f"{round(r['actual_sku']/r['max_sku']*100)}%" if pd.notna(r["max_sku"]) and r["max_sku"] > 0 else "N/A", axis=1)
    result["units_pct"] = result.apply(
        lambda r: f"{round(r['actual_units']/r['max_units']*100)}%" if pd.notna(r["max_units"]) and r["max_units"] > 0 else "N/A", axis=1)
    return result[[
        "category","max_sku","actual_sku","sku_diff","sku_pct",
        "max_units","actual_units","units_diff","units_pct"
    ]].rename(columns={
        "category":"Category","max_sku":"Max SKU","actual_sku":"Actual SKU",
        "sku_diff":"SKU Available","sku_pct":"SKU Used %",
        "max_units":"Max Units","actual_units":"Actual Units",
        "units_diff":"Units Available","units_pct":"Units Used %"
    })

# ============================================================
# 顶部导航
# ============================================================
col1, col2, col3 = st.columns([4, 1, 1])
with col1:
    st.title("🛍️ PopStop Retail Dashboard")
with col2:
    st.write(f"👤 {st.session_state.user}")
with col3:
    if st.button("Logout", key="logout_btn"):
        st.session_state.logged_in = False
        st.rerun()

# 显示数据最后更新时间（UTC+12 NZT）
try:
    import datetime as _dt
    parquet_path  = os.path.join(DATA_DIR, "output_company.parquet")
    mod_timestamp = os.path.getmtime(parquet_path)
    nzt_offset    = _dt.timezone(_dt.timedelta(hours=12))
    update_dt     = _dt.datetime.fromtimestamp(mod_timestamp, tz=_dt.timezone.utc).astimezone(nzt_offset)
    update_str    = update_dt.strftime("%d %b %Y %H:%M")
    st.caption(f"📅 Data last updated: {update_str} (NZT)")
except Exception:
    pass

st.markdown("---")
role       = st.session_state.role
user_store = st.session_state.store

# ============================================================
# Company View
# ============================================================
def show_company_view():
    st.subheader("🏢 Company Sales Overview")
    co_df     = load_company(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_company.parquet")))
    store_df  = load_store(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_store.parquet")))
    limits_df = load_limits(_mtime=_get_mtime(LIMITS_FILE))

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        search = st.text_input("🔍 Search SKU / Product", "", key="co_search")
    with col2:
        cats = ["All"] + sorted(co_df["category"].dropna().unique().tolist())
        cat = st.selectbox("Category", cats, key="co_cat")
    with col3:
        brands = ["All"] + sorted(co_df["brand"].dropna().unique().tolist())
        brand = st.selectbox("Brand", brands, key="co_brand")
    with col4:
        suppliers = ["All"] + sorted(co_df["supplier"].dropna().unique().tolist())
        supplier = st.selectbox("Supplier", suppliers, key="co_sup")

    filtered_co = co_df.copy()
    if search:
        filtered_co = filtered_co[
            filtered_co["sku"].astype(str).str.contains(search, case=False, na=False) |
            filtered_co["product"].astype(str).str.contains(search, case=False, na=False) |
            filtered_co["supplier_code"].astype(str).str.contains(search, case=False, na=False)
        ]
    if cat != "All":
        filtered_co = filtered_co[filtered_co["category"] == cat]
    if brand != "All":
        filtered_co = filtered_co[filtered_co["brand"] == brand]
    if supplier != "All":
        filtered_co = filtered_co[filtered_co["supplier"] == supplier]

    # Sort by STORE_TOT descending
    if "STORE_TOT" in filtered_co.columns:
        filtered_co = filtered_co.sort_values("STORE_TOT", ascending=False)

    filtered_store = store_df[store_df["sku"].isin(filtered_co["sku"])]

    # Load preorders, in-transit, and pending-invoice data
    po_mtime  = _get_mtime(os.path.join(DATA_DIR, "preorders.json"))
    preorders = load_preorders(_mtime=po_mtime)
    sh_mtime  = _get_mtime(os.path.join(DATA_DIR, "shipment_db.json"))
    shipments = load_shipment_db(_mtime=sh_mtime)
    transit_map = build_transit_map(shipments)  # {supplier_code_upper: "QG:12 RICC:6"}
    pend_mtime = _get_mtime(os.path.join(DATA_DIR, "pending_orders.json"))
    pending_map = load_pending_orders_map(_mtime=pend_mtime)

    filtered_co = filtered_co.copy()
    filtered_co["Pre-order"] = filtered_co["supplier_code"].map(
        lambda sc: preorders.get(str(sc), {}).get("total_qty", 0) if pd.notna(sc) and str(sc).strip() else 0
    ).fillna(0).astype(int)
    filtered_co["In Transit"] = filtered_co["supplier_code"].map(
        lambda sc: transit_map.get(str(sc).upper(), "") if pd.notna(sc) and str(sc).strip() else ""
    ).fillna("")
    filtered_co["Pending Invoice"] = filtered_co["supplier_code"].map(
        lambda sc: (
            f"{pending_map[str(sc).upper()]['remaining_qty']} "
            f"(✈️{pending_map[str(sc).upper()]['remaining_air']}/"
            f"🚢{pending_map[str(sc).upper()]['remaining_sea']} · "
            f"{'OVERDUE ' if pending_map[str(sc).upper()]['overdue'] else ''}"
            f"{pending_map[str(sc).upper()]['days_pending']}d)"
        ) if pd.notna(sc) and str(sc).strip() and str(sc).upper() in pending_map else ""
    ).fillna("")

    # Reorder columns: move Pre-order, In Transit, Pending Invoice after TOT
    if "TOT" in filtered_co.columns:
        cols = filtered_co.columns.tolist()
        extra = [c for c in ["Pre-order","In Transit","Pending Invoice"] if c in cols]
        other = [c for c in cols if c not in extra]
        if "TOT" in other:
            idx = other.index("TOT") + 1
            new_order = other[:idx] + extra + other[idx:]
            filtered_co = filtered_co[new_order]

    po_count   = (filtered_co["Pre-order"] > 0).sum()
    tr_count   = (filtered_co["In Transit"] != "").sum()
    pend_count = (filtered_co["Pending Invoice"] != "").sum()
    info_parts = []
    if po_count > 0: info_parts.append(f"🟡 {po_count} with pre-orders")
    if tr_count > 0: info_parts.append(f"🟠 {tr_count} in transit")
    if pend_count > 0: info_parts.append(f"⏳ {pend_count} pending invoice")
    st.write(f"**{len(filtered_co):,} products**" + (" | " + " | ".join(info_parts) if info_parts else ""))

    # ── 隐藏列 ───────────────────────────────────────────────
    HIDE_COLS = {"id","tag","supply_price","retail_price","inventory_cost",
                 "retail_value","brand","created","first_sale",
                 "days_since_created","days_since_first_sale"}

    # ── 列分组（用于底色）────────────────────────────────────
    SOH_COLS    = {"DUN","PAP","QG","RICC","RICH","SP","TR","Office",
                   "WH1","AKLWH","TOT","Pre-order","In Transit","Pending Invoice"}
    L1W_COLS    = {"DUN_L1W","PAP_L1W","QG_L1W","RICC_L1W","RICH_L1W",
                   "SP_L1W","TR_L1W","Office_L1W","STORE_TOT"}
    METRIC_COLS = {"items_sold_lifetime","CY26","avg_weekly","WOH",
                   "L6W","L3W","L1W","Company_Status","Action"}

    # ── 整数列 ───────────────────────────────────────────────
    INT_COLS = {"DUN","PAP","QG","RICC","RICH","SP","TR","Office","WH1","AKLWH","TOT",
                "DUN_L1W","PAP_L1W","QG_L1W","RICC_L1W","RICH_L1W","SP_L1W",
                "TR_L1W","Office_L1W","STORE_TOT","Pre-order",
                "items_sold_lifetime","CY26","L6W","L3W","L1W",
                "days_since_last_sold","days_since_last_received"}

    PAGE_SIZE   = 1000
    total_pages = max(1, (len(filtered_co) - 1) // PAGE_SIZE + 1)
    page = st.number_input(f"Page (1-{total_pages})", min_value=1,
                           max_value=total_pages, value=1, key="co_page") if total_pages > 1 else 1
    start   = (page - 1) * PAGE_SIZE
    page_df = filtered_co.iloc[start:start+PAGE_SIZE].copy()

    # 去掉隐藏列
    show_cols = [c for c in page_df.columns if c not in HIDE_COLS]
    page_df   = page_df[show_cols]

    # 整数格式 - 所有数量列显示整数无小数
    for c in INT_COLS:
        if c in page_df.columns:
            page_df[c] = pd.to_numeric(page_df[c], errors="coerce").fillna(0).astype(int)

    # avg_weekly, WOH 保留1位小数
    for c in ["avg_weekly","WOH"]:
        if c in page_df.columns:
            page_df[c] = pd.to_numeric(page_df[c], errors="coerce").round(1)

    # 周列也转整数
    week_cols = [c for c in page_df.columns if c.startswith("2026-W") or c.startswith("2025-W")]
    for c in week_cols:
        page_df[c] = pd.to_numeric(page_df[c], errors="coerce").fillna(0).astype(int).replace(0, None)

    def highlight_co(row):
        styles = []
        for col in row.index:
            has_transit = bool(row.get("In Transit",""))
            has_po      = row.get("Pre-order",0) > 0
            # Row highlight for in-transit/pre-order
            if has_transit:
                base = "background-color:#FFF3E0"  # 全行淡橙
            elif has_po:
                base = "background-color:#FFFDE7"  # 全行淡黄
            else:
                base = ""
            # Column group highlight (override base for specific cols)
            if col in SOH_COLS:
                styles.append("background-color:#E3F2FD")   # 淡蓝
            elif col in L1W_COLS:
                styles.append("background-color:#E8F5E9")   # 淡绿
            elif col in METRIC_COLS:
                styles.append("background-color:#F3E5F5")   # 淡紫
            else:
                styles.append(base)
        return styles

    st.dataframe(
        page_df.style.apply(highlight_co, axis=1),
        width="stretch", height=600,
        column_config={
            "sku":          st.column_config.TextColumn("SKU",     pinned=True, width=130),
            "product":      st.column_config.TextColumn("Product", pinned=True, width=280),
            "avg_weekly":   st.column_config.NumberColumn("Avg/Wk", format="%.1f"),
            "WOH":          st.column_config.NumberColumn("WOH",    format="%.1f"),
            "supply_price": st.column_config.NumberColumn("Cost",   format="$%.2f"),
            "retail_price": st.column_config.NumberColumn("Retail", format="$%.2f"),
        }
    )

    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        if st.button("📥 Download Company Data", key="co_prep"):
            with st.spinner("Generating Excel..."):
                excel_data = generate_excel(filtered_co, pd.DataFrame(), limits_df, store_name=None)
            st.download_button("⬇️ Download Company",
                excel_data, file_name="PopStop_Company.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="co_download")
    with col_dl2:
        st.info("💡 To download store reports, go to **Store View** → select 'All' stores.")

# ============================================================
# Store View
# ============================================================
def show_store_view(store_filter=None):
    st.subheader("🏪 Store Sales Overview")
    co_df     = load_company(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_company.parquet")))
    store_df  = load_store(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_store.parquet")))
    limits_df = load_limits(_mtime=_get_mtime(LIMITS_FILE))

    if store_filter:
        filtered_store = store_df[store_df["store"] == store_filter].copy()
        st.info(f"**Store: {store_filter}**")
        current_store = store_filter
    else:
        stores   = ["All"] + sorted(store_df["store"].dropna().unique().tolist())
        selected = st.selectbox("Select Store", stores, key="st_store")
        current_store  = selected
        filtered_store = store_df[store_df["store"] == selected].copy() if selected != "All" else store_df.copy()

    col1, col2, col3 = st.columns(3)
    with col1:
        search = st.text_input("🔍 Search SKU / Product", "", key="st_search")
    with col2:
        cats = ["All"] + sorted(filtered_store["category"].dropna().unique().tolist())
        cat = st.selectbox("Category", cats, key="st_cat")
    with col3:
        brands = ["All"] + sorted(filtered_store["brand"].dropna().unique().tolist())
        brand = st.selectbox("Brand", brands, key="st_brand")

    if search:
        filtered_store = filtered_store[
            filtered_store["sku"].astype(str).str.contains(search, case=False, na=False) |
            filtered_store["product"].astype(str).str.contains(search, case=False, na=False) |
            filtered_store["supplier_code"].astype(str).str.contains(search, case=False, na=False)
        ]
    if cat != "All":
        filtered_store = filtered_store[filtered_store["category"] == cat]
    if brand != "All":
        filtered_store = filtered_store[filtered_store["brand"] == brand]

    filtered_co = co_df[co_df["sku"].isin(filtered_store["sku"])]

    # Load preorders, in-transit, and pending-invoice data
    po_mtime  = _get_mtime(os.path.join(DATA_DIR, "preorders.json"))
    preorders = load_preorders(_mtime=po_mtime)
    sh_mtime  = _get_mtime(os.path.join(DATA_DIR, "shipment_db.json"))
    shipments = load_shipment_db(_mtime=sh_mtime)
    transit_map = build_transit_map(shipments)
    pend_mtime = _get_mtime(os.path.join(DATA_DIR, "pending_orders.json"))
    pending_map = load_pending_orders_map(_mtime=pend_mtime)

    filtered_store = filtered_store.copy()
    filtered_store["Pre-order"] = filtered_store["supplier_code"].map(
        lambda sc: preorders.get(str(sc), {}).get("total_qty", 0) if pd.notna(sc) and str(sc).strip() else 0
    ).fillna(0).astype(int)
    filtered_store["In Transit"] = filtered_store["supplier_code"].map(
        lambda sc: transit_map.get(str(sc).upper(), "") if pd.notna(sc) and str(sc).strip() else ""
    ).fillna("")
    filtered_store["Pending Invoice"] = filtered_store["supplier_code"].map(
        lambda sc: (
            f"{pending_map[str(sc).upper()]['remaining_qty']} "
            f"(✈️{pending_map[str(sc).upper()]['remaining_air']}/"
            f"🚢{pending_map[str(sc).upper()]['remaining_sea']} · "
            f"{'OVERDUE ' if pending_map[str(sc).upper()]['overdue'] else ''}"
            f"{pending_map[str(sc).upper()]['days_pending']}d)"
        ) if pd.notna(sc) and str(sc).strip() and str(sc).upper() in pending_map else ""
    ).fillna("")

    po_count   = (filtered_store["Pre-order"] > 0).sum()
    tr_count   = (filtered_store["In Transit"] != "").sum()
    pend_count = (filtered_store["Pending Invoice"] != "").sum()
    info_parts = []
    if po_count > 0: info_parts.append(f"🟡 {po_count} with pre-orders")
    if tr_count > 0: info_parts.append(f"🟠 {tr_count} in transit")
    if pend_count > 0: info_parts.append(f"⏳ {pend_count} pending invoice")
    st.write(f"**{len(filtered_store):,} products**" + (" | " + " | ".join(info_parts) if info_parts else ""))

    st.markdown(
        "🟦 **L1W** &nbsp;&nbsp; 🟩 **L3W** &nbsp;&nbsp; 🟧 **L6W** &nbsp;&nbsp; 🟪 **WOH** &nbsp;&nbsp;&nbsp;"
        "🟢 **WOH store better** &nbsp;&nbsp; 🔴 **WOH store worse** &nbsp;&nbsp; "
        "🟡 **Pre-order** &nbsp;&nbsp; 🟠 **In Transit**"
    )

    # ── 隐藏列（同 Company View）────────────────────────────
    HIDE_COLS_ST = {"id","tag","supply_price","retail_price","inventory_cost",
                    "retail_value","brand","created","first_sale",
                    "days_since_created","days_since_first_sale"}

    # ── 列分组底色 ────────────────────────────────────────────
    SOH_COLS_ST    = {"soh","SOH_co","Pre-order","In Transit"}
    L1W_COLS_ST    = {"L1W","L1W_co","L3W","L3W_co","L6W","L6W_co"}
    WOH_COLS_ST    = {"WOH_store","WOH_co","avg_weekly","avg_weekly_co"}
    METRIC_COLS_ST = {"CY26","items_sold","items_sold_lifetime_co",
                      "Store_Status","Company_Status","Action"}

    # ── 整数列 ────────────────────────────────────────────────
    INT_COLS_ST = {"soh","SOH_co","L1W","L1W_co","L3W","L3W_co","L6W","L6W_co",
                   "CY26","items_sold","items_sold_lifetime_co","Pre-order",
                   "days_since_last_sold","days_since_last_received"}

    PAGE_SIZE   = 500
    total_pages = max(1, (len(filtered_store) - 1) // PAGE_SIZE + 1)
    page = st.number_input(f"Page (1-{total_pages})", min_value=1,
                           max_value=total_pages, value=1, key="st_page") if total_pages > 1 else 1
    start      = (page - 1) * PAGE_SIZE
    display_df = filtered_store.iloc[start:start+PAGE_SIZE].reset_index(drop=True).copy()

    # 去掉隐藏列
    show_cols_st = [c for c in display_df.columns if c not in HIDE_COLS_ST]
    display_df   = display_df[show_cols_st]

    # 整数格式
    for c in INT_COLS_ST:
        if c in display_df.columns:
            display_df[c] = pd.to_numeric(display_df[c], errors="coerce").fillna(0).astype(int)

    # avg_weekly, WOH 保留1位小数
    for c in ["avg_weekly","avg_weekly_co","WOH_store","WOH_co"]:
        if c in display_df.columns:
            display_df[c] = pd.to_numeric(display_df[c], errors="coerce").round(1)

    # 周列转整数
    week_cols_st = [c for c in display_df.columns if c.startswith("2026-W") or c.startswith("2025-W")]
    for c in week_cols_st:
        display_df[c] = pd.to_numeric(display_df[c], errors="coerce").fillna(0).astype(int).replace(0, None)

    def highlight_st(row):
        styles = []
        for col in row.index:
            has_transit = bool(row.get("In Transit",""))
            has_po      = row.get("Pre-order",0) > 0
            if has_transit:
                base = "background-color:#FFF3E0"
            elif has_po:
                base = "background-color:#FFFDE7"
            else:
                base = ""
            if col in SOH_COLS_ST:
                styles.append("background-color:#E3F2FD")
            elif col in L1W_COLS_ST:
                styles.append("background-color:#E8F5E9")
            elif col in WOH_COLS_ST:
                styles.append("background-color:#F3E5F5")
            elif col in METRIC_COLS_ST:
                styles.append("background-color:#EDE7F6")
            else:
                styles.append(base)
        return styles

    st.dataframe(
        display_df.style.apply(highlight_st, axis=1),
        width="stretch", height=600,
        column_config={
            "sku":          st.column_config.TextColumn("SKU",     pinned=True, width=130),
            "product":      st.column_config.TextColumn("Product", pinned=True, width=280),
            "avg_weekly":   st.column_config.NumberColumn("Avg/Wk",    format="%.1f"),
            "avg_weekly_co":st.column_config.NumberColumn("Avg/Wk Co", format="%.1f"),
            "WOH_store":    st.column_config.NumberColumn("WOH St",    format="%.1f"),
            "WOH_co":       st.column_config.NumberColumn("WOH Co",    format="%.1f"),
        }
    )

    if current_store != "All":
        if st.button(f"📥 Download {current_store} Report", key="st_prep"):
            with st.spinner("Generating Excel..."):
                excel_data = generate_excel(
                    filtered_co, filtered_store, limits_df,
                    store_name=current_store)
            st.download_button("⬇️ Click to Download", excel_data,
                file_name=f"PopStop_{current_store}_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="st_download")
    else:
        st.markdown("---")
        STORE_LIST = ["Dunedin","Papanui","Queensgate","Riccarton",
                      "Richmond","Sylvia Park","Te Rapa","Office"]
        STORE_NAME_VARIANTS = {
            "Sylvia Park": ["Sylvia Park","SylviaPark","Sylvia_Park"],
            "Te Rapa":     ["Te Rapa","TeRapa","Te_Rapa","Terapa"],
        }
        dl_sel = st.multiselect("Select stores to download",
                                STORE_LIST, default=STORE_LIST, key="st_all_sel")
        if st.button("📥 Download Selected Stores", key="st_all_prep"):
            with st.spinner("Generating store reports..."):
                store_df_dl = load_store(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_store.parquet")))
                store_col   = "store" if "store" in store_df_dl.columns else "outlet"
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    for store_nm in dl_sel:
                        variants = STORE_NAME_VARIANTS.get(store_nm, [store_nm])
                        sdf = pd.DataFrame()
                        for v in variants:
                            sdf = store_df_dl[store_df_dl[store_col] == v]
                            if not sdf.empty:
                                break
                        if sdf.empty:
                            continue
                        sdf.to_excel(writer, index=False, sheet_name=store_nm[:28])
                        _format_sheet(writer.sheets[store_nm[:28]], sdf, is_store=True)
                del store_df_dl
            st.download_button("⬇️ Download All Store Reports", buf.getvalue(),
                file_name="PopStop_Store_Reports.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="st_all_download")

# ============================================================
# Store Limits View
# ============================================================
def show_limits_view(store_filter=None):
    st.subheader("📊 Store SKU & Unit Limits")
    store_df  = load_store(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_store.parquet")))
    limits_df = load_limits(_mtime=_get_mtime(LIMITS_FILE))

    if limits_df.empty:
        st.warning("No limits data found. Please check store_limits.csv")
        return

    available_stores = limits_df["store"].unique().tolist()

    if store_filter and store_filter in available_stores:
        current_store = store_filter
    elif store_filter and store_filter not in available_stores:
        st.info(f"No limits data set up for {store_filter} yet. Coming soon!")
        return
    else:
        current_store = st.selectbox(
            "Select Store", available_stores, key="lim_store"
        ) if available_stores else None

    if not current_store:
        st.info("No limits data available yet.")
        return

    result = calc_limits_comparison(store_df, limits_df, current_store)
    if result is None or result.empty:
        st.info(f"No limits data for {current_store} yet.")
        return

    st.write(f"**Store: {current_store}**")
    st.markdown("🔴 **Over limit** &nbsp;&nbsp; 🟡 **≥80% used** &nbsp;&nbsp; 🟢 **Under limit**")

    def color_limits(df):
        styled = pd.DataFrame("", index=df.index, columns=df.columns)
        for idx in df.index:
            try:
                max_s = float(df.at[idx,"Max SKU"])
                act_s = float(df.at[idx,"Actual SKU"])
                if max_s > 0:
                    pct = act_s / max_s
                    if pct > 1:
                        styled.at[idx,"Actual SKU"]    = "background-color: #FFCDD2; font-weight: bold"
                        styled.at[idx,"SKU Available"] = "background-color: #FFCDD2"
                    elif pct >= 0.8:
                        styled.at[idx,"Actual SKU"]    = "background-color: #FFF9C4"
                        styled.at[idx,"SKU Available"] = "background-color: #FFF9C4"
                    else:
                        styled.at[idx,"Actual SKU"]    = "background-color: #C8E6C9"
            except:
                pass
            try:
                max_u = float(df.at[idx,"Max Units"])
                act_u = float(df.at[idx,"Actual Units"])
                if max_u > 0:
                    pct = act_u / max_u
                    if pct > 1:
                        styled.at[idx,"Actual Units"]    = "background-color: #FFCDD2; font-weight: bold"
                        styled.at[idx,"Units Available"] = "background-color: #FFCDD2"
                    elif pct >= 0.8:
                        styled.at[idx,"Actual Units"]    = "background-color: #FFF9C4"
                        styled.at[idx,"Units Available"] = "background-color: #FFF9C4"
                    else:
                        styled.at[idx,"Actual Units"]    = "background-color: #C8E6C9"
            except:
                pass
        return styled

    try:
        styled = result.style.apply(color_limits, axis=None)
        st.dataframe(styled, width="stretch", height=600)
    except:
        st.dataframe(result, width="stretch", height=600)

    col1, col2, col3, col4 = st.columns(4)
    over_sku   = len(result[result["Actual SKU"]   > result["Max SKU"]])
    over_units = len(result[result["Actual Units"] > result["Max Units"]])
    with col1:
        st.metric("Categories Over SKU Limit",   over_sku,   delta_color="inverse")
    with col2:
        st.metric("Categories Over Units Limit", over_units, delta_color="inverse")
    with col3:
        st.metric("Total SKU Used",
                  f"{result['Actual SKU'].sum()} / {result['Max SKU'].sum()}")
    with col4:
        st.metric("Total Units Used",
                  f"{result['Actual Units'].sum()} / {result['Max Units'].sum()}")

    st.markdown("---")
    dl_col1, dl_col2 = st.columns(2)

    def build_limits_export(store_name, limits_df_in, shipments_lim, sku_cat_lim,
                            weekly_data_lim, co_df_lim, ACTIVE_ST):
        from collections import defaultdict

        # Get result for this store
        store_df_tmp = load_store(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_store.parquet")))
        result_tmp = calc_limits_comparison(store_df_tmp, limits_df_in, store_name)
        if result_tmp is None or result_tmp.empty:
            return []

        # Batches
        batch_tmp = defaultdict(lambda: {"edd":"","cat_skus":defaultdict(set),"cat_qty":defaultdict(int)})
        for inv_no, shp in shipments_lim.items():
            if shp.get("store") != store_name: continue
            if shp.get("tnl_status") not in ACTIVE_ST: continue
            ref = shp.get("customer_ref") or inv_no
            edd = shp.get("edd","") or shp.get("eta","")
            if not batch_tmp[ref]["edd"] or (edd and edd < batch_tmp[ref]["edd"]):
                batch_tmp[ref]["edd"] = edd
            for item in shp.get("items",[]):
                sku = item.get("sku","").strip().upper()
                qty = item.get("quantity",0) or 0
                cat = sku_cat_lim.get(sku,"Unknown")
                batch_tmp[ref]["cat_skus"][cat].add(sku)
                batch_tmp[ref]["cat_qty"][cat] += qty
        sorted_batches_tmp = sorted(batch_tmp.items(), key=lambda x: x[1]["edd"] or "9999")

        # Avg from weekly data
        cat_avg_u, cat_avg_s = {}, {}
        if weekly_data_lim and store_name in weekly_data_lim.get("cat_stores", {}):
            sr  = weekly_data_lim["cat_stores"][store_name]
            nw  = len(weekly_data_lim.get("weeks", [1]))
            if "SALES_U" in sr:
                su = sr["SALES_U"].copy(); gc = su.columns[0]
                wkc = [c for c in su.columns if c.startswith("2026-W")]
                su["_t"] = su[wkc].sum(axis=1)
                for _, r in su.iterrows(): cat_avg_u[r[gc]] = round(r["_t"]/max(nw,1),1)
            if "SKU" in sr:
                sk = sr["SKU"].copy(); gc2 = sk.columns[0]
                wks = [c for c in sk.columns if c.startswith("2026-W")]
                sk["_t"] = sk[wks].sum(axis=1)
                for _, r in sk.iterrows(): cat_avg_s[r[gc2]] = round(r["_t"]/max(nw,1),1)

        rows_out = []
        for _, row in result_tmp.iterrows():
            cat = row["Category"]
            r = {
                "Store":         store_name,
                "Category":      cat,
                "Max SKU":       int(row["Max SKU"])   if pd.notna(row["Max SKU"])   and str(row["Max SKU"])   != "None" else "",
                "Max Units":     int(row["Max Units"]) if pd.notna(row["Max Units"]) and str(row["Max Units"]) != "None" else "",
                "Actual SKU":    int(row["Actual SKU"]),
                "Actual Units":  int(row["Actual Units"]),
                "Avg WK SKU":    cat_avg_s.get(cat, 0),
                "Avg WK Units":  cat_avg_u.get(cat, 0),
            }
            for ref, batch in sorted_batches_tmp:
                edd = batch["edd"] or "TBC"
                r[f"{ref} ({edd}) SKU"] = len(batch["cat_skus"].get(cat, set()))
                r[f"{ref} ({edd}) QTY"] = batch["cat_qty"].get(cat, 0)
            rows_out.append(r)
        return rows_out

    def load_common_data():
        from collections import defaultdict
        db_path = os.path.join(DATA_DIR, "shipment_db.json")
        shipments_lim = {}
        if os.path.exists(db_path):
            with open(db_path) as f:
                db = json.load(f)
            shipments_lim = db.get("shipments", {})
        co_df_lim = load_company(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_company.parquet")))
        sku_cat_lim = dict(zip(
            co_df_lim["supplier_code"].astype(str).str.upper(),
            co_df_lim["category"].fillna("Unknown")
        ))
        ACTIVE_ST = {"Pending","Waiting for Shipping","In Transit",
                     "Documents Received","In Customs","Arrived NZ","Customs Cleared"}
        weekly_data_lim = load_weekly(_mtime=_get_mtime(os.path.join(DATA_DIR,"weekly_reports.pkl")))
        return shipments_lim, sku_cat_lim, ACTIVE_ST, co_df_lim, weekly_data_lim

    def format_limits_sheet(ws):
        hdr_fill = PatternFill("solid", fgColor="37474F")
        hdr_font = Font(color="FFFFFF", bold=True)
        for ci in range(1, ws.max_column+1):
            ws.cell(1, ci).fill = hdr_fill
            ws.cell(1, ci).font = hdr_font
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 22
        for ci in range(3, ws.max_column+1):
            ws.column_dimensions[get_column_letter(ci)].width = 15
        ws.freeze_panes = "B2"

    # ── Button 1: Single store ────────────────────────────────
    with dl_col1:
        if st.button(f"📥 Download {current_store} Report", key="lim_dl_prep"):
            with st.spinner("Generating..."):
                shipments_lim, sku_cat_lim, ACTIVE_ST, co_df_lim, weekly_data_lim = load_common_data()
                rows = build_limits_export(current_store, limits_df, shipments_lim,
                                           sku_cat_lim, weekly_data_lim, co_df_lim, ACTIVE_ST)
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    df_out = pd.DataFrame(rows)
                    # Remove Store column for single store
                    if "Store" in df_out.columns:
                        df_out = df_out.drop(columns=["Store"])
                    df_out.to_excel(writer, index=False, sheet_name=f"{current_store[:28]}_Limits")
                    format_limits_sheet(writer.sheets[f"{current_store[:28]}_Limits"])
            st.download_button(f"⬇️ Download {current_store} Limits",
                buf.getvalue(), f"PopStop_{current_store}_Limits.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="lim_dl")

    # ── Button 2: All stores in one sheet ────────────────────
    with dl_col2:
        if st.button("📥 Download ALL Stores (One Sheet)", key="lim_all_prep"):
            with st.spinner("Generating all stores..."):
                shipments_lim, sku_cat_lim, ACTIVE_ST, co_df_lim, weekly_data_lim = load_common_data()
                STORE_LIST = ["Dunedin","Papanui","Queensgate","Riccarton",
                              "Richmond","Sylvia Park","Te Rapa","Office"]
                all_rows = []

                # Company summary row per category
                store_df_all = load_store(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_store.parquet")))
                all_cats = sorted(limits_df["category"].dropna().unique().tolist())
                nw = len(weekly_data_lim.get("weeks", [1])) if weekly_data_lim else 1

                for cat in all_cats:
                    # Company actuals
                    co_cat_df = co_df_lim[co_df_lim["category"] == cat] if not co_df_lim.empty else pd.DataFrame()
                    co_actual_sku   = int(co_cat_df["sku"].nunique()) if not co_cat_df.empty else 0
                    co_actual_units = int(co_cat_df["TOT"].sum()) if not co_cat_df.empty and "TOT" in co_cat_df.columns else 0

                    # Company avg from weekly
                    co_avg_u, co_avg_s = 0, 0
                    if weekly_data_lim and "cat_company" in weekly_data_lim:
                        cc = weekly_data_lim["cat_company"]
                        if "SALES_U" in cc:
                            su = cc["SALES_U"]; gc = su.columns[0]
                            wkc = [c for c in su.columns if c.startswith("2026-W")]
                            row_co = su[su[gc] == cat]
                            if not row_co.empty:
                                co_avg_u = round(float(row_co[wkc].sum(axis=1).values[0]) / max(nw,1), 1)
                        if "SKU" in cc:
                            sk = cc["SKU"]; gc2 = sk.columns[0]
                            wks = [c for c in sk.columns if c.startswith("2026-W")]
                            row_sk = sk[sk[gc2] == cat]
                            if not row_sk.empty:
                                co_avg_s = round(float(row_sk[wks].sum(axis=1).values[0]) / max(nw,1), 1)

                    all_rows.append({
                        "Store": "🏢 PopStop (ALL)",
                        "Category": cat,
                        "Max SKU": "", "Max Units": "",
                        "Actual SKU": co_actual_sku,
                        "Actual Units": co_actual_units,
                        "Avg WK SKU": co_avg_s,
                        "Avg WK Units": co_avg_u,
                    })

                # Each store
                for store_nm in STORE_LIST:
                    if store_nm not in limits_df["store"].values:
                        continue
                    rows_s = build_limits_export(store_nm, limits_df, shipments_lim,
                                                  sku_cat_lim, weekly_data_lim, co_df_lim, ACTIVE_ST)
                    all_rows.extend(rows_s)

                buf = BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    df_all = pd.DataFrame(all_rows)
                    df_all.to_excel(writer, index=False, sheet_name="All_Stores_Limits")
                    ws_all = writer.sheets["All_Stores_Limits"]
                    format_limits_sheet(ws_all)

                    # Highlight company summary rows
                    co_fill = PatternFill("solid", fgColor="E8EAF6")
                    co_font = Font(bold=True)
                    for ri in range(2, ws_all.max_row+1):
                        cell_val = ws_all.cell(ri, 1).value
                        if cell_val and "PopStop (ALL)" in str(cell_val):
                            for ci in range(1, ws_all.max_column+1):
                                ws_all.cell(ri, ci).fill = co_fill
                                ws_all.cell(ri, ci).font = co_font

            st.download_button("⬇️ Download All Stores Limits",
                buf.getvalue(), "PopStop_AllStores_Limits.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="lim_all_dl")

# ============================================================
# Weekly Reports View
# ============================================================
def show_weekly_view():
    st.subheader("📈 Weekly Sales Reports")
    weekly_data = load_weekly(_mtime=_get_mtime(os.path.join(DATA_DIR,"weekly_reports.pkl")))

    if weekly_data is None:
        st.warning("No weekly report data found. Please run process_data.py first.")
        return

    weeks           = weekly_data["weeks"]
    weeks_sorted    = sorted(weeks, reverse=True)
    cat_company     = weekly_data["cat_company"]
    cat_stores      = weekly_data["cat_stores"]
    sup_company     = weekly_data["sup_company"]
    brand_company   = weekly_data.get("brand_company", {})
    brand_by_cat_co = weekly_data.get("brand_by_cat_company", {})
    brand_stores    = weekly_data.get("brand_stores", {})
    brand_by_cat_st = weekly_data.get("brand_by_cat_stores", {})
    tag_company     = weekly_data.get("tag_company", {})
    tag_by_cat_co   = weekly_data.get("tag_by_cat_company", {})
    tag_stores      = weekly_data.get("tag_stores", {})

    st.write(f"**2026 — {len(weeks)} weeks ({weeks_sorted[-1]} to {weeks_sorted[0]})**")

    wr_tab1, wr_tab2, wr_tab3, wr_tab4 = st.tabs([
        "📊 Weekly Sales Report",
        "🏥 Category Health",
        "📋 Category Weekly Report",
        "🏆 Category Store Comparison"
    ])

    # ── Load company parquet for inventory data ───────────────
    co_df_wr = load_company(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_company.parquet")))

    with wr_tab2:
        st.markdown("### 🏥 Category Inventory Health")
        st.caption("SOH vs full-year average weekly sales")

        if co_df_wr is not None and not co_df_wr.empty:
            # Full-year avg weekly = total all-week sales / number of weeks
            n_weeks = len(weeks_sorted)
            last_wk = weeks_sorted[0]

            # Get total sales per SKU across all weeks from cat_company REV & SALES_U
            # Use SALES_U from weekly_data cat_company
            rev_df    = cat_company.get("REV",    {})
            sales_df  = cat_company.get("SALES_U", {})

            # Build category health from co_df
            cat_health = co_df_wr.groupby("category").agg(
                SKUs        = ("sku",            "count"),
                Total_SOH   = ("TOT",            "sum"),
                Inv_Cost    = ("inventory_cost",  "sum"),
                Retail_Val  = ("retail_value",    "sum"),
            ).reset_index()
            cat_health.rename(columns={"category":"Category"}, inplace=True)

            # Full-year avg weekly from weekly_data SALES_U
            if isinstance(sales_df, pd.DataFrame):
                wk_cols = [c for c in sales_df.columns if c.startswith("2026-W")]
                sales_df["fy_avg"] = sales_df[wk_cols].sum(axis=1) / len(wk_cols)
                cat_health = cat_health.merge(
                    sales_df[["category","fy_avg"]].rename(columns={"category":"Category"}),
                    on="Category", how="left"
                )
            else:
                # Build from cat_company SALES_U dict
                fy_rows = []
                if "SALES_U" in cat_company:
                    su = cat_company["SALES_U"]
                    wk_cols = [c for c in su.columns if c.startswith("2026-W")]
                    su_copy = su.copy()
                    su_copy["fy_avg"] = su_copy[wk_cols].sum(axis=1) / max(len(wk_cols),1)
                    fy_rows = su_copy[["category","fy_avg"]].rename(columns={"category":"Category"})
                if fy_rows is not None and len(fy_rows) > 0:
                    cat_health = cat_health.merge(fy_rows, on="Category", how="left")
                else:
                    cat_health["fy_avg"] = 0

            if "fy_avg" not in cat_health.columns:
                cat_health["fy_avg"] = 0

            # Also add L6W avg from co_df
            cat_l6w = co_df_wr.groupby("category")["L6W"].sum().reset_index()
            cat_l6w.columns = ["Category","L6W_total"]
            cat_l6w["L6W_avg"] = (cat_l6w["L6W_total"] / 6).round(1)
            cat_health = cat_health.merge(cat_l6w, on="Category", how="left")

            cat_health["FY Avg/Week"]  = cat_health["fy_avg"].round(1)
            cat_health["L6W Avg/Week"] = cat_health["L6W_avg"].fillna(0).round(1)
            cat_health["WOH (FY)"]     = (cat_health["Total_SOH"] / cat_health["fy_avg"].replace(0, float("nan"))).round(1)
            cat_health["WOH (L6W)"]    = (cat_health["Total_SOH"] / cat_health["L6W_avg"].replace(0, float("nan"))).round(1)
            cat_health["Inv Cost"]     = cat_health["Inv_Cost"].round(0).fillna(0).astype(int)
            cat_health["Retail Value"] = cat_health["Retail_Val"].round(0).fillna(0).astype(int)

            display_health = cat_health[[
                "Category","SKUs","Total_SOH","FY Avg/Week","L6W Avg/Week",
                "WOH (FY)","WOH (L6W)","Inv Cost","Retail Value"
            ]].rename(columns={"Total_SOH":"Total SOH"}).sort_values("Total SOH", ascending=False)

            def color_woh(val):
                try:
                    v = float(val)
                    if v < 4:   return "background-color:#FFCDD2"
                    if v < 8:   return "background-color:#FFF9C4"
                    if v > 26:  return "background-color:#E3F2FD"
                    return "background-color:#C8E6C9"
                except:
                    return ""

            st.dataframe(
                display_health.style.map(color_woh, subset=["WOH (FY)","WOH (L6W)"]),
                width="stretch", hide_index=True, height=600
            )
            st.caption("🔴 WOH<4 critical | 🟡 WOH<8 low | 🟢 WOH 8-26 healthy | 🔵 WOH>26 overstock")

            # Download button
            if st.button("📥 Download Category Health", key="cat_health_prep"):
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    display_health.to_excel(writer, index=False, sheet_name="Category Health")
                st.download_button("⬇️ Download", buf.getvalue(),
                    "PopStop_Category_Health.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="cat_health_dl")
        else:
            st.info("No company data available.")

    with wr_tab3:
        st.markdown("### 📋 Category Weekly Performance")

        if "SALES_U" in cat_company and "REV" in cat_company:
            cats_available = sorted(cat_company["REV"]["category"].tolist()
                if "category" in cat_company["REV"].columns
                else cat_company["REV"].iloc[:,0].tolist())

            selected_cat_wr = st.selectbox("Select Category", ["All"] + cats_available, key="cat_wr_sel")

            # Build comparison table: latest week vs previous week
            last_wk = weeks_sorted[0]
            prev_wk = weeks_sorted[1] if len(weeks_sorted) > 1 else None

            def get_cat_metric(metric_key, cat=None, wk=None):
                if metric_key not in cat_company: return 0
                df = cat_company[metric_key]
                grp_col = df.columns[0]
                if wk not in df.columns: return 0
                if cat and cat != "All":
                    row = df[df[grp_col] == cat]
                    return float(row[wk].values[0]) if not row.empty else 0
                return float(df[wk].sum())

            # Summary cards
            st.markdown(f"#### {selected_cat_wr} — {last_wk} vs {prev_wk}")
            metrics = [
                ("Revenue",    "REV",    "${:,.0f}",  False),
                ("Items Sold", "SALES_U","{:,.0f}",   False),
                ("GP",         "GP",     "${:,.0f}",  False),
                ("Margin",     "MARGIN", "{:.1%}",    True),
            ]
            cols = st.columns(4)
            for i, (label, key, fmt, is_pct_m) in enumerate(metrics):
                cur = get_cat_metric(key, selected_cat_wr, last_wk)
                prv = get_cat_metric(key, selected_cat_wr, prev_wk) if prev_wk else 0
                delta = cur - prv
                if is_pct_m:
                    cols[i].metric(label, f"{cur:.1%}", f"{delta:+.1%}")
                else:
                    cols[i].metric(label, fmt.format(cur), f"{delta:+,.0f}")

            st.markdown("---")

            # Weekly trend table for this category
            st.markdown("##### Weekly Trend")
            trend_rows = []
            for wk in weeks_sorted:
                rev  = get_cat_metric("REV",    selected_cat_wr, wk)
                gp   = get_cat_metric("GP",     selected_cat_wr, wk)
                units= get_cat_metric("SALES_U",selected_cat_wr, wk)
                margin = (gp/rev) if rev > 0 else 0
                trend_rows.append({
                    "Week":       wk,
                    "Revenue":    f"${rev:,.0f}",
                    "Items Sold": f"{units:,.0f}",
                    "GP":         f"${gp:,.0f}",
                    "Margin":     f"{margin:.1%}",
                })
            trend_df = pd.DataFrame(trend_rows)
            st.dataframe(trend_df, width="stretch", hide_index=True, height=500)

            # Top 5 SKUs this week
            if selected_cat_wr != "All" and co_df_wr is not None:
                st.markdown(f"##### Top SKUs — {last_wk}")
                top_skus = co_df_wr[co_df_wr["category"] == selected_cat_wr].copy()
                if last_wk in top_skus.columns:
                    top_skus = top_skus[top_skus[last_wk] > 0][
                        ["product","supplier_code",last_wk,"L6W","TOT","WOH","Company_Status","Action"]
                    ].sort_values(last_wk, ascending=False).head(10)
                    top_skus.columns = ["Product","SKU",f"Units ({last_wk})","L6W","SOH","WOH","Status","Action"]
                    st.dataframe(top_skus, width="stretch", hide_index=True)

            # Download Category Weekly Report
            if st.button("📥 Download Category Report", key="cat_wr_dl_prep"):
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    trend_df.to_excel(writer, index=False, sheet_name=f"{selected_cat_wr[:25]}_Trend")
                st.download_button("⬇️ Download", buf.getvalue(),
                    f"PopStop_{selected_cat_wr}_Weekly.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="cat_wr_dl")
        else:
            st.info("No weekly category data available.")

    with wr_tab4:
        st.markdown("### 🏆 Category Store Comparison")
        st.caption("Compare each category's performance across all stores — revenue, store share (% of store total), and company share (% of company's category total).")

        if "REV" not in cat_company or not cat_stores:
            st.info("No data available.")
        else:
            rev_co  = cat_company["REV"].copy()
            grp_col = rev_co.columns[0]
            all_cats = sorted(rev_co[grp_col].dropna().tolist())
            VALID_STORES = ["Dunedin","Papanui","Queensgate","Riccarton",
                            "Richmond","Sylvia Park","Te Rapa","Office"]
            store_list = [s for s in VALID_STORES if s in cat_stores]

            # Controls row
            ctrl1, ctrl2, ctrl3 = st.columns([2, 2, 2])
            with ctrl1:
                sel_cat = st.selectbox("Category", all_cats, key="csc_cat")
            with ctrl2:
                n_weeks_show = st.slider("Weeks to show", 4, min(len(weeks_sorted), 20),
                                         min(8, len(weeks_sorted)), key="csc_weeks")
            with ctrl3:
                view_mode = st.radio("View", ["📊 Table", "🌡️ Heatmap", "📈 Chart"],
                                     horizontal=True, key="csc_view")

            display_weeks = weeks_sorted[:n_weeks_show]

            # ── Build data matrix ────────────────────────────────────
            # For each store × week: rev, store_share (% of store total), co_share (% of co cat total)
            def get_rev(report, cat, wk):
                if "REV" not in report: return 0.0
                df = report["REV"]
                gc = df.columns[0]
                if wk not in df.columns: return 0.0
                row = df[df[gc] == cat]
                return float(row[wk].values[0]) if not row.empty else 0.0

            def get_store_total(report, wk):
                if "REV" not in report: return 0.0
                df = report["REV"]
                if wk not in df.columns: return 0.0
                return float(df[wk].sum())

            def get_co_cat_rev(wk):
                return get_rev(cat_company, sel_cat, wk)

            rows = []
            for store in store_list:
                report = cat_stores.get(store, {})
                row = {"Store": store}
                for wk in display_weeks:
                    rev        = get_rev(report, sel_cat, wk)
                    st_total   = get_store_total(report, wk)
                    co_cat_rev = get_co_cat_rev(wk)
                    st_share   = rev / st_total   if st_total   > 0 else 0.0
                    co_share   = rev / co_cat_rev if co_cat_rev > 0 else 0.0
                    row[f"{wk}_rev"]      = rev
                    row[f"{wk}_st_share"] = st_share
                    row[f"{wk}_co_share"] = co_share
                rows.append(row)

            matrix_df = pd.DataFrame(rows)

            # ── Rank each store per week by revenue ──────────────────
            for wk in display_weeks:
                col_rev = f"{wk}_rev"
                if col_rev in matrix_df.columns:
                    matrix_df[f"{wk}_rank"] = matrix_df[col_rev].rank(ascending=False, method="min").astype(int)

            # ── TABLE VIEW ───────────────────────────────────────────
            if view_mode == "📊 Table":
                table_rows = []
                for _, r in matrix_df.iterrows():
                    trow = {"Store": r["Store"]}
                    for wk in display_weeks:
                        rev      = r.get(f"{wk}_rev", 0)
                        st_share = r.get(f"{wk}_st_share", 0)
                        co_share = r.get(f"{wk}_co_share", 0)
                        rank     = r.get(f"{wk}_rank", "-")
                        wk_short = wk.replace("2026-","")
                        trow[f"{wk_short} Rev"]    = f"${rev:,.0f}" if rev else "-"
                        trow[f"{wk_short} Store%"] = f"{st_share:.1%}" if st_share else "-"
                        trow[f"{wk_short} Co%"]    = f"{co_share:.1%}" if co_share else "-"
                        trow[f"{wk_short} Rank"]   = f"#{rank}" if rev else "-"
                    table_rows.append(trow)

                table_df = pd.DataFrame(table_rows)
                st.dataframe(table_df, width="stretch", hide_index=True,
                             height=min(400, (len(store_list)+1)*38))

            # ── HEATMAP VIEW ─────────────────────────────────────────
            elif view_mode == "🌡️ Heatmap":
                heat_metric = st.radio(
                    "Colour by",
                    ["Store% (share of store)", "Co% (share of company cat)", "Revenue ($)"],
                    horizontal=True, key="csc_heat_metric"
                )
                suffix_map = {
                    "Store% (share of store)":       "_st_share",
                    "Co% (share of company cat)":    "_co_share",
                    "Revenue ($)":                   "_rev",
                }
                suffix = suffix_map[heat_metric]
                is_pct = suffix != "_rev"

                heat_data = matrix_df[["Store"]].copy()
                for wk in display_weeks:
                    wk_short = wk.replace("2026-", "")
                    heat_data[wk_short] = matrix_df[f"{wk}{suffix}"]

                heat_data = heat_data.set_index("Store")

                def color_cell(val):
                    try:
                        v = float(val)
                        if v == 0: return "background-color:#F5F5F5; color:#AAAAAA"
                        if is_pct:
                            # Green gradient for pct
                            if v >= 0.30:  return "background-color:#1B5E20; color:white"
                            elif v >= 0.20: return "background-color:#2E7D32; color:white"
                            elif v >= 0.15: return "background-color:#43A047; color:white"
                            elif v >= 0.10: return "background-color:#81C784"
                            elif v >= 0.05: return "background-color:#C8E6C9"
                            else:           return "background-color:#F1F8E9"
                        else:
                            mx = heat_data.values.max()
                            if mx == 0: return ""
                            ratio = v / mx
                            if ratio >= 0.80:  return "background-color:#1565C0; color:white"
                            elif ratio >= 0.60: return "background-color:#1976D2; color:white"
                            elif ratio >= 0.40: return "background-color:#42A5F5"
                            elif ratio >= 0.20: return "background-color:#90CAF9"
                            else:               return "background-color:#E3F2FD"
                    except:
                        return ""

                def fmt_cell(val):
                    try:
                        v = float(val)
                        if v == 0: return "-"
                        return f"{v:.1%}" if is_pct else f"${v:,.0f}"
                    except:
                        return str(val)

                styled = heat_data.style.map(color_cell).format(fmt_cell)
                st.dataframe(styled, width="stretch",
                             height=min(400, (len(store_list)+1)*38))

                # Also show rank table below heatmap
                st.caption("**Rank by Revenue** (within category, per week)")
                rank_data = matrix_df[["Store"]].copy()
                for wk in display_weeks:
                    wk_short = wk.replace("2026-","")
                    rank_data[wk_short] = matrix_df.apply(
                        lambda r: f"#{int(r[f'{wk}_rank'])}" if r.get(f"{wk}_rev",0) > 0 else "-", axis=1)
                st.dataframe(rank_data.set_index("Store"), width="stretch",
                             height=min(350, (len(store_list)+1)*38))

            # ── CHART VIEW ───────────────────────────────────────────
            elif view_mode == "📈 Chart":
                chart_metric = st.radio(
                    "Metric",
                    ["Store% (share of store)", "Co% (share of company cat)", "Revenue ($)"],
                    horizontal=True, key="csc_chart_metric"
                )
                suffix_map2 = {
                    "Store% (share of store)":       "_st_share",
                    "Co% (share of company cat)":    "_co_share",
                    "Revenue ($)":                   "_rev",
                }
                suffix2 = suffix_map2[chart_metric]
                is_pct2 = suffix2 != "_rev"

                chart_df = pd.DataFrame(index=[wk.replace("2026-","") for wk in reversed(display_weeks)])
                for store in store_list:
                    row = matrix_df[matrix_df["Store"] == store].iloc[0]
                    chart_df[store] = [row.get(f"{wk}{suffix2}", 0) for wk in reversed(display_weeks)]

                if is_pct2:
                    chart_df = chart_df * 100
                    st.caption(f"**{sel_cat}** — {chart_metric} (%) per week per store")
                else:
                    st.caption(f"**{sel_cat}** — Revenue ($) per week per store")

                st.line_chart(chart_df, use_container_width=True, height=350)

            # ── DOWNLOAD ─────────────────────────────────────────────
            st.markdown("---")
            if st.button("📥 Download Store Comparison", key="csc_dl_prep"):
                with st.spinner("Building Excel..."):
                    buf = BytesIO()
                    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                        # Sheet 1: Full data table
                        export_rows = []
                        for _, r in matrix_df.iterrows():
                            for wk in weeks_sorted:  # all weeks in download
                                rev      = r.get(f"{wk}_rev", 0)
                                st_share = r.get(f"{wk}_st_share", 0)
                                co_share = r.get(f"{wk}_co_share", 0)
                                rank     = r.get(f"{wk}_rank", None)
                                export_rows.append({
                                    "Store":    r["Store"],
                                    "Week":     wk,
                                    "Revenue":  round(rev, 2),
                                    "Store%":   round(st_share, 4),
                                    "Co%":      round(co_share, 4),
                                    "Rank":     int(rank) if pd.notna(rank) and rev > 0 else None,
                                })
                        export_df = pd.DataFrame(export_rows)
                        export_df.to_excel(writer, index=False, sheet_name=f"{sel_cat[:25]}_Data")

                        # Sheet 2: Store% heatmap pivot
                        wb2 = writer.book
                        ws2 = wb2.create_sheet(f"{sel_cat[:20]}_StoreShare")
                        ws2.cell(1,1,"Store \\ Week")
                        all_wks_rev = weeks_sorted
                        for ci, wk in enumerate(all_wks_rev, 2):
                            ws2.cell(1, ci, wk.replace("2026-",""))
                            ws2.cell(1, ci).font = Font(bold=True)
                        from openpyxl.formatting.rule import ColorScaleRule
                        for ri, store in enumerate(store_list, 2):
                            ws2.cell(ri, 1, store)
                            row = matrix_df[matrix_df["Store"] == store]
                            if row.empty: continue
                            row = row.iloc[0]
                            for ci, wk in enumerate(all_wks_rev, 2):
                                val = row.get(f"{wk}_st_share", 0)
                                c = ws2.cell(ri, ci, round(val, 4) if val else None)
                                c.number_format = "0.0%"
                        # Color scale
                        last_col = get_column_letter(len(all_wks_rev)+1)
                        last_row = len(store_list)+1
                        ws2.conditional_formatting.add(
                            f"B2:{last_col}{last_row}",
                            ColorScaleRule(
                                start_type="num", start_value=0, start_color="F5F5F5",
                                mid_type="percentile", mid_value=50, mid_color="81C784",
                                end_type="percentile", end_value=100, end_color="1B5E20"
                            )
                        )
                        ws2.column_dimensions["A"].width = 18
                        ws2.freeze_panes = "B2"

                        # Sheet 3: Co% heatmap pivot
                        ws3 = wb2.create_sheet(f"{sel_cat[:20]}_CoShare")
                        ws3.cell(1,1,"Store \\ Week")
                        for ci, wk in enumerate(all_wks_rev, 2):
                            ws3.cell(1, ci, wk.replace("2026-",""))
                            ws3.cell(1, ci).font = Font(bold=True)
                        for ri, store in enumerate(store_list, 2):
                            ws3.cell(ri, 1, store)
                            row = matrix_df[matrix_df["Store"] == store]
                            if row.empty: continue
                            row = row.iloc[0]
                            for ci, wk in enumerate(all_wks_rev, 2):
                                val = row.get(f"{wk}_co_share", 0)
                                c = ws3.cell(ri, ci, round(val, 4) if val else None)
                                c.number_format = "0.0%"
                        ws3.conditional_formatting.add(
                            f"B2:{last_col}{last_row}",
                            ColorScaleRule(
                                start_type="num", start_value=0, start_color="F5F5F5",
                                mid_type="percentile", mid_value=50, mid_color="90CAF9",
                                end_type="percentile", end_value=100, end_color="1565C0"
                            )
                        )
                        ws3.column_dimensions["A"].width = 18
                        ws3.freeze_panes = "B2"

                    st.download_button(
                        f"⬇️ Download {sel_cat} Store Comparison",
                        buf.getvalue(),
                        file_name=f"PopStop_{sel_cat}_StoreComparison.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="csc_dl"
                    )

    with wr_tab1:
        # 报表类型选择
        report_options = [
            "Category — Company",
            "Category — By Store",
            "Supplier — Company",
            "Brand/IP — Company (All Categories)",
            "Brand/IP — Company (By Category)",
            "Brand/IP — By Store (All Categories)",
            "Brand/IP — By Store (By Category)",
            "Tag — Company (All Categories)",
            "Tag — Company (By Category)",
            "Tag — By Store",
        ]

        col1, col2 = st.columns(2)
        with col1:
            report_type = st.selectbox("Report Type", report_options, key="wr_type")
        with col2:
            metric = st.selectbox(
                "Metric",
                ["REV $ (Revenue)", "GP $ (Gross Profit)",
                 "MARGIN % (Margin)", "SALES U (Units)", "SALES SKU (# SKUs Sold)", "REV % of Category"],
                key="wr_metric"
            )

        metric_map = {
            "REV $ (Revenue)":        "REV",
            "GP $ (Gross Profit)":    "GP",
            "MARGIN % (Margin)":      "MARGIN",
            "SALES U (Units)":        "SALES_U",
            "SALES SKU (# SKUs Sold)":"SKU",
            "REV % of Category":      "REV_PCT",
        }
        metric_key = metric_map[metric]
        is_pct     = metric_key in ["MARGIN","REV_PCT"]

        report       = None
        need_store   = "By Store" in report_type
        need_cat     = "By Category" in report_type
        selected_store = None
        selected_cat   = None

        if need_store:
            store_options  = list(cat_stores.keys())
            selected_store = st.selectbox("Select Store", store_options, key="wr_store")

        if need_cat:
            cat_options  = sorted(brand_by_cat_co.keys())
            selected_cat = st.selectbox("Select Category", cat_options, key="wr_cat")

        if report_type == "Category — Company":
            report = cat_company
        elif report_type == "Category — By Store" and selected_store:
            report = cat_stores.get(selected_store)
        elif report_type == "Supplier — Company":
            report = sup_company
        elif report_type == "Brand/IP — Company (All Categories)":
            report = brand_company
        elif report_type == "Brand/IP — Company (By Category)" and selected_cat:
            report = brand_by_cat_co.get(selected_cat)
        elif report_type == "Brand/IP — By Store (All Categories)" and selected_store:
            report = brand_stores.get(selected_store)
        elif report_type == "Brand/IP — By Store (By Category)" and selected_store and selected_cat:
            report = brand_by_cat_st.get(selected_store, {}).get(selected_cat)
        elif report_type == "Tag — Company (All Categories)":
            report = tag_company
        elif report_type == "Tag — Company (By Category)" and selected_cat:
            report = tag_by_cat_co.get(selected_cat)
        elif report_type == "Tag — By Store" and selected_store:
            report = tag_stores.get(selected_store)

        if report and metric_key in report:
            df        = report[metric_key].copy()
            group_col = df.columns[0]
            wk_cols_avail = [w for w in weeks_sorted if w in df.columns]
            l4w_cols = wk_cols_avail[:4]
            l8w_cols = wk_cols_avail[:min(8, len(wk_cols_avail))]

            display_df = df.copy()
            for wk in wk_cols_avail:
                if is_pct:
                    display_df[wk] = display_df[wk].apply(
                        lambda x: f"{x*100:.1f}%" if pd.notna(x) and x != 0 else "0%")
                else:
                    display_df[wk] = display_df[wk].apply(
                        lambda x: f"{x:,.0f}" if pd.notna(x) else "0")

            # Add L4W Avg and L8W Avg columns
            if len(l4w_cols) >= 2:
                for i, row in df.iterrows():
                    v4 = pd.to_numeric(pd.Series([row.get(w,0) for w in l4w_cols]), errors="coerce").mean()
                    v8 = pd.to_numeric(pd.Series([row.get(w,0) for w in l8w_cols]), errors="coerce").mean()
                    display_df.at[i, "L4W Avg"] = f"{v4*100:.1f}%" if is_pct else f"{v4:,.1f}"
                    display_df.at[i, "L8W Avg"] = f"{v8*100:.1f}%" if is_pct else f"{v8:,.1f}"

            # Build avg row across all categories
            avg_row = {group_col: "📊 AVG / Week"}
            for wk in wk_cols_avail:
                vals = pd.to_numeric(df[wk], errors="coerce").dropna()
                v = vals.mean() if len(vals) > 0 else 0
                avg_row[wk] = f"{v*100:.1f}%" if is_pct else f"{v:,.1f}"
            if len(l4w_cols) >= 2:
                v4_all = pd.to_numeric(df[l4w_cols].values.flatten(), errors="coerce").mean()
                v8_all = pd.to_numeric(df[l8w_cols].values.flatten(), errors="coerce").mean()
                avg_row["L4W Avg"] = f"{v4_all*100:.1f}%" if is_pct else f"{v4_all:,.1f}"
                avg_row["L8W Avg"] = f"{v8_all*100:.1f}%" if is_pct else f"{v8_all:,.1f}"

            ordered_cols = [group_col] + (["L4W Avg","L8W Avg"] if "L4W Avg" in avg_row else []) + [w for w in weeks_sorted if w in display_df.columns]
            display_df = display_df[[c for c in ordered_cols if c in display_df.columns]]

            avg_df_row = pd.DataFrame([{c: avg_row.get(c,"") for c in display_df.columns}])
            display_df = pd.concat([display_df, avg_df_row], ignore_index=True)

            st.dataframe(display_df, width="stretch", height=520)
        else:
            st.info("Please select report type and filters above.")

        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("📥 Download Full Weekly Report", key="wr_prep"):
                with st.spinner("Generating Excel..."):
                    excel_data = generate_weekly_excel(weekly_data, weeks_sorted)
                st.download_button(
                    "⬇️ Download Full Weekly Report",
                    excel_data,
                    file_name="PopStop_Weekly_Report_2026.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="wr_download"
                )
        with col2:
            if selected_cat:
                scope = "store" if selected_store else "company"
                if st.button(f"📥 Download {selected_cat} Detail", key="wr_cat_prep"):
                    with st.spinner("Generating Excel..."):
                        excel_data = generate_category_detail_excel(
                            weekly_data, selected_cat, weeks_sorted,
                            scope=scope, store_name=selected_store)
                    st.download_button(
                        f"⬇️ Download {selected_cat} Detail",
                        excel_data,
                        file_name=f"PopStop_{selected_cat}_Detail.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="wr_cat_download"
                    )

# ============================================================
# Last Week Summary
# ============================================================
def show_last_week_summary():
    st.subheader("📋 Last Week Sales Summary")
    weekly_data = load_weekly(_mtime=_get_mtime(os.path.join(DATA_DIR,"weekly_reports.pkl")))

    if weekly_data is None:
        st.warning("No weekly report data found. Please run process_data.py first.")
        return

    weeks        = weekly_data["weeks"]
    weeks_sorted = sorted(weeks, reverse=True)
    last_week    = weeks_sorted[0]

    cat_company = weekly_data["cat_company"]
    cat_stores  = weekly_data["cat_stores"]

    st.info(f"📅 Showing data for **{last_week}**")

    # ── 从 cat_company 提取公司总计 ──────────────────────────────
    def get_company_total(metric_key):
        if metric_key not in cat_company:
            return None
        df = cat_company[metric_key]
        if last_week not in df.columns:
            return None
        return df[last_week].sum()

    co_rev   = get_company_total("REV")
    co_gp    = get_company_total("GP")
    co_units = get_company_total("SALES_U")
    co_margin = (co_gp / co_rev) if co_rev and co_gp and co_rev != 0 else None

    # ── 从 cat_stores 提取各门店 ─────────────────────────────────
    # ── 从 cat_stores 提取各门店（key是outlet名称）─────────────────
    # cat_stores的key来自outlet列：Dunedin, Papanui, Queensgate, Riccarton,
    # Richmond, Sylvia Park, Te Rapa, Office
    VALID_STORES_DISPLAY = [
        ("Dunedin",     "Dunedin"),
        ("Papanui",     "Papanui"),
        ("Queensgate",  "Queensgate"),
        ("Riccarton",   "Riccarton"),
        ("Richmond",    "Richmond"),
        ("Sylvia Park", "Sylvia Park"),
        ("Te Rapa",     "Te Rapa"),
        ("Office",      "Office"),
    ]

    store_rows = []
    for store_key, store_label in VALID_STORES_DISPLAY:
        report = cat_stores.get(store_key)
        if report is None:
            continue
        def _sum(mk):
            if mk not in report:
                return None
            df = report[mk]
            if last_week not in df.columns:
                return None
            return df[last_week].sum()
        rev   = _sum("REV")
        gp    = _sum("GP")
        units = _sum("SALES_U")
        if rev is None and gp is None and units is None:
            continue
        margin = (gp / rev) if rev and gp and rev != 0 else None
        store_rows.append({
            "Store":       store_label,
            "Revenue":     rev or 0,
            "Items Sold":  int(units or 0),
            "GP":          gp or 0,
            "Margin":      margin,
        })

    # ── 公司汇总卡片 ──────────────────────────────────────────────
    st.markdown("### 🏢 Company Total")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Revenue",    f"${co_rev:,.2f}"    if co_rev    is not None else "N/A")
    with c2:
        st.metric("Items Sold", f"{int(co_units):,}" if co_units  is not None else "N/A")
    with c3:
        st.metric("GP",         f"${co_gp:,.2f}"     if co_gp     is not None else "N/A")
    with c4:
        st.metric("Margin",     f"{co_margin*100:.1f}%" if co_margin is not None else "N/A")

    # ── 门店明细表 ────────────────────────────────────────────────
    st.markdown("### 🏪 Store Breakdown")
    if store_rows:
        summary_df = pd.DataFrame(store_rows)

        def fmt_summary(df):
            styled = pd.DataFrame("", index=df.index, columns=df.columns)
            return styled

        display_df = summary_df.copy()
        display_df["Revenue"]    = display_df["Revenue"].apply(lambda x: f"${x:,.2f}")
        display_df["Items Sold"] = display_df["Items Sold"].apply(lambda x: f"{x:,}")
        display_df["GP"]         = display_df["GP"].apply(lambda x: f"${x:,.2f}")
        display_df["Margin"]     = display_df["Margin"].apply(
            lambda x: f"{x*100:.1f}%" if x is not None else "N/A")

        st.dataframe(display_df, width="stretch", hide_index=True)
    else:
        st.info("No store data available for this week.")

    # ── 下载Excel ─────────────────────────────────────────────────
    st.markdown("---")
    if st.button("📥 Download Last Week Summary", key="lws_prep"):
        with st.spinner("Generating Excel..."):
            excel_data = generate_last_week_excel(
                last_week, co_rev, co_units, co_gp, co_margin, store_rows)
        st.download_button(
            "⬇️ Download Last Week Summary",
            excel_data,
            file_name=f"PopStop_Summary_{last_week}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="lws_download"
        )


def generate_last_week_excel(last_week, co_rev, co_units, co_gp, co_margin, store_rows):
    buffer = BytesIO()
    wb     = openpyxl.Workbook()
    ws     = wb.active
    ws.title = f"Summary_{last_week}"

    border      = _get_border()
    header_fill = PatternFill("solid", fgColor=XL_COLORS["header_bg"])
    header_font = Font(color=XL_COLORS["header_font"], bold=True)
    alt_fill    = PatternFill("solid", fgColor=XL_COLORS["row_alt"])
    norm_fill   = PatternFill("solid", fgColor="FFFFFF")
    total_fill  = PatternFill("solid", fgColor="37474F")
    total_font  = Font(color="FFFFFF", bold=True)

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"PopStop Weekly Sales Summary — {last_week}"
    title_cell.font  = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", fgColor="D32F2F")
    title_cell.border = border

    # Header row
    headers = ["Store", "Revenue", "Items Sold", "GP", "Margin"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.border = border

    # Company Total row
    row = 3
    data = [
        "PopStop (Total)",
        co_rev or 0,
        int(co_units or 0),
        co_gp or 0,
        co_margin,
    ]
    fmts = [None, '#,##0.00', '#,##0', '#,##0.00', '0.00%']
    for ci, (val, fmt) in enumerate(zip(data, fmts), start=1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.fill   = total_fill
        cell.font   = total_font
        cell.border = border
        if fmt:
            cell.number_format = fmt

    # Store rows
    for ri, sr in enumerate(store_rows):
        row += 1
        fill = alt_fill if ri % 2 == 0 else norm_fill
        vals = [
            sr["Store"],
            sr["Revenue"],
            sr["Items Sold"],
            sr["GP"],
            sr["Margin"] if sr["Margin"] is not None else 0,
        ]
        for ci, (val, fmt) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill   = fill
            cell.border = border
            if fmt:
                cell.number_format = fmt

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col in ["B","C","D","E"]:
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = "A3"

    wb.save(buffer)
    return buffer.getvalue()


# ============================================================
# Shipment Hub
# ============================================================
def show_store_requests():
    st.subheader("🛒 Store Requests")
    st.caption("Products with negative SOH (pre-orders/backorders), with in-transit and pre-order information.")

    co_df  = load_company(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_company.parquet")))
    weekly = load_weekly(_mtime=_get_mtime(os.path.join(DATA_DIR,"weekly_reports.pkl")))
    po_mtime  = _get_mtime(os.path.join(DATA_DIR,"preorders.json"))
    preorders = load_preorders(_mtime=po_mtime)

    db_path   = os.path.join(DATA_DIR,"shipment_db.json")
    shipments = {}
    if os.path.exists(db_path):
        with open(db_path) as f:
            shipments = json.load(f).get("shipments",{})

    # Store SOH columns in co_df
    STORE_COLS = {
        "DUN":"DUN","PAP":"PAP","QG":"QG","RICC":"RICC",
        "RICH":"RICH","SP":"SP","TR":"TR","Office":"Office","WH1":"WH1"
    }
    soh_store_cols = [c for c in co_df.columns if c in STORE_COLS.values()]

    # Find negative SOH products
    if not soh_store_cols:
        st.warning("No store SOH columns found in company data.")
        return

    neg_mask = (co_df[soh_store_cols] < 0).any(axis=1)
    neg_df   = co_df[neg_mask].copy()

    if neg_df.empty:
        st.success("✅ No backorders or pre-orders found this week.")
        return

    st.info(f"Found **{len(neg_df)}** products with backorders/pre-orders")

    # Get current week label from today's date (not from pkl)
    from datetime import date
    today = date.today()
    iso_week = today.isocalendar()
    wk_short = f"W{iso_week[1]:02d}"          # e.g. "W23"
    last_wk  = f"{iso_week[0]}-{wk_short}"    # e.g. "2026-W23"

    # Build in-transit lookup by supplier_code
    transit_by_sku = {}  # sku -> {qty, edd}
    ACTIVE_ST = {"Pending","Waiting for Shipping","In Transit",
                 "Documents Received","In Customs","Arrived NZ","Customs Cleared"}
    for inv_no, shp in shipments.items():
        if shp.get("tnl_status") not in ACTIVE_ST: continue
        edd = shp.get("edd","") or shp.get("eta","")
        for item in shp.get("items",[]):
            sku = item.get("sku","").strip().upper()
            qty = item.get("quantity",0) or 0
            if sku not in transit_by_sku:
                transit_by_sku[sku] = {"qty":0,"edd":""}
            transit_by_sku[sku]["qty"] += qty
            if not transit_by_sku[sku]["edd"] or (edd and edd < transit_by_sku[sku]["edd"]):
                transit_by_sku[sku]["edd"] = edd

    # Get L1W sales per store from weekly data
    l1w_by_sku_store = {}  # (sku, store_abbrev) -> qty
    if weekly:
        cat_stores = weekly.get("cat_stores",{})
        # Use company level L1W from co_df directly
        pass

    # Build output rows
    rows_out = []
    for _, row in neg_df.iterrows():
        sku          = str(row.get("supplier_code","")).strip().upper()
        sku_id       = str(row.get("sku","")).strip()
        product      = row.get("product","")
        category     = row.get("category","")
        supplier     = row.get("supplier","")
        retail_price = row.get("retail_price","")

        # SOH per store
        soh_vals = {c: int(row.get(c,0) or 0) for c in soh_store_cols}
        tot_soh  = sum(soh_vals.values())

        # L1W per store from co_df
        l1w_cols = {c.replace("DUN","DUN").replace("PAP","PAP"): c
                    for c in co_df.columns if c.startswith("L1W_") or c == "L1W"}

        # In transit
        transit_info = transit_by_sku.get(sku, transit_by_sku.get(sku_id.upper(), {}))
        transit_qty  = transit_info.get("qty", 0)
        transit_edd  = transit_info.get("edd", "")

        # Pre-order (IKON)
        po_info  = preorders.get(sku, preorders.get(sku_id, {}))
        po_qty   = po_info.get("total_qty", 0) if po_info else 0
        po_eta   = po_info.get("eta","") if po_info else ""

        r = {
            "Product":      product,
            "SKU":          sku_id,
            "Category":     category,
            "Supplier":     supplier,
            "Price (NZD)":  retail_price,
        }
        # SOH columns
        for c in soh_store_cols:
            r[c] = soh_vals[c] if soh_vals[c] != 0 else ""
        r["TOT SOH"] = tot_soh

        # In Transit
        r["In Transit QTY"] = transit_qty if transit_qty > 0 else ""
        r["In Transit EDD"] = transit_edd if transit_qty > 0 else ""

        # Pre-order
        r["Pre-order QTY"] = po_qty if po_qty > 0 else ""
        r["Pre-order ETA"] = po_eta if po_qty > 0 else ""

        # Coverage note
        shortfall = abs(tot_soh)
        if transit_qty >= shortfall:
            r["Coverage"] = "✅ Covered by transit"
        elif transit_qty > 0:
            r["Coverage"] = f"⚠️ Partial ({transit_qty}/{shortfall})"
        elif po_qty > 0:
            r["Coverage"] = f"🔵 Pre-order only"
        else:
            r["Coverage"] = "❌ No coverage"

        rows_out.append(r)

    result_df = pd.DataFrame(rows_out)

    # Display filters
    fc1, fc2 = st.columns(2)
    cats = sorted(result_df["Category"].dropna().unique())
    sel_cat = fc1.multiselect("Category", cats, key="sr_cat")
    cov_opts = ["✅ Covered by transit","⚠️ Partial","🔵 Pre-order only","❌ No coverage"]
    sel_cov = fc2.multiselect("Coverage", cov_opts, key="sr_cov")

    disp_df = result_df.copy()
    if sel_cat: disp_df = disp_df[disp_df["Category"].isin(sel_cat)]
    if sel_cov: disp_df = disp_df[disp_df["Coverage"].str.contains("|".join([c[:3] for c in sel_cov]))]

    st.dataframe(disp_df, width="stretch", hide_index=True, height=500)
    st.caption(f"{len(disp_df)} products | Week: {wk_short}")

    st.markdown("---")
    if st.button(f"📥 Download WK{wk_short} Store Requests", key="sr_dl_prep"):
        with st.spinner("Generating Excel..."):
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                result_df.to_excel(writer, index=False, sheet_name=f"WK{wk_short}_Store_Requests")
                ws = writer.sheets[f"WK{wk_short}_Store_Requests"]

                # Header formatting
                hdr_fill = PatternFill("solid", fgColor="37474F")
                hdr_font = Font(color="FFFFFF", bold=True)
                # Section colors
                soh_fill     = PatternFill("solid", fgColor="E3F2FD")  # blue tint - SOH
                transit_fill = PatternFill("solid", fgColor="E8F5E9")  # green tint - transit
                po_fill      = PatternFill("solid", fgColor="FFF3E0")  # orange tint - pre-order

                headers = [ws.cell(1, i).value for i in range(1, ws.max_column+1)]
                for ci, h in enumerate(headers, 1):
                    cell = ws.cell(1, ci)
                    cell.font = hdr_font
                    if h in soh_store_cols or h == "TOT SOH":
                        cell.fill = PatternFill("solid", fgColor="1565C0")
                    elif h and "Transit" in str(h):
                        cell.fill = PatternFill("solid", fgColor="2E7D32")
                    elif h and "Pre-order" in str(h):
                        cell.fill = PatternFill("solid", fgColor="E65100")
                    elif h == "Coverage":
                        cell.fill = PatternFill("solid", fgColor="4A148C")
                    else:
                        cell.fill = hdr_fill

                # Column widths
                ws.column_dimensions["A"].width = 50  # Product
                ws.column_dimensions["B"].width = 16  # SKU
                ws.column_dimensions["C"].width = 18  # Category
                ws.column_dimensions["D"].width = 16  # Supplier
                for ci in range(5, ws.max_column+1):
                    ws.column_dimensions[get_column_letter(ci)].width = 14

                # Highlight negative SOH cells in red
                from openpyxl.formatting.rule import CellIsRule
                soh_col_letters = []
                for ci, h in enumerate(headers, 1):
                    if h in soh_store_cols or h == "TOT SOH":
                        col_letter = get_column_letter(ci)
                        soh_col_letters.append(col_letter)
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            CellIsRule(operator="lessThan", formula=["0"],
                                      fill=PatternFill("solid", fgColor="FFCDD2"))
                        )

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"

        st.download_button(
            f"⬇️ Download WK{wk_short} Store Requests",
            buf.getvalue(),
            file_name=f"WK{wk_short}_2026_Store_Requests.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="sr_dl"
        )


# ============================================================
# Shipment Hub
# ============================================================
def show_inventory_recommendations():
    st.markdown("### 🔄 Inventory Recommendations — Transfer & Replenishment")
    st.caption(
        "Auto-generated suggestions based on per-store sales velocity and stock levels. "
        "Run `generate_inventory_recommendations.py` locally to refresh this data."
    )

    rec_path = os.path.join(DATA_DIR, "inventory_recommendations.json")
    if not os.path.exists(rec_path):
        st.info(
            "No inventory_recommendations.json found yet. Run "
            "`generate_inventory_recommendations.py` on your local machine to generate this data."
        )
        return

    with open(rec_path) as f:
        rec_data = json.load(f)

    gen_at = rec_data.get("generated_at", "")
    if gen_at:
        try:
            gen_dt = datetime.strptime(gen_at[:19], "%Y-%m-%dT%H:%M:%S")
            st.caption(f"📅 Last generated: {gen_dt.strftime('%d %b %Y %H:%M')}")
        except Exception:
            pass

    replen    = rec_data.get("replenishment", [])
    transfers = rec_data.get("transfers", [])
    clearance = rec_data.get("clearance_candidates", [])

    c1, c2, c3 = st.columns(3)
    c1.metric("📦 Replenish Suggestions", len(replen))
    c2.metric("🔁 Transfer Suggestions", len(transfers))
    c3.metric("🏷️ Clearance Candidates", len(clearance))

    st.markdown("---")

    sub_replen, sub_transfer, sub_clearance = st.tabs([
        "📦 Warehouse → Store", "🔁 Store ↔ Store", "🏷️ Clearance Candidates"
    ])

    with sub_replen:
        st.caption(
            "Stores selling this SKU well (or already owing pre-orders to customers) "
            "but running low, where the AKL/Office warehouse has spare stock available."
        )
        if not replen:
            st.success("✅ No replenishment needed right now.")
        else:
            backorder_n = sum(1 for r in replen if r.get("backorder_units", 0) > 0)
            if backorder_n:
                st.warning(f"⚠️ {backorder_n} of these involve units already on backorder/pre-order — customers are already waiting.")

            store_opts = sorted(set(r["to_store"] for r in replen))
            sel_stores = st.multiselect("Filter by store", store_opts, key="replen_store_filter")
            show_backorder_only = st.checkbox("⚠️ Show backorders only", key="replen_backorder_only")

            rows = [r for r in replen if not sel_stores or r["to_store"] in sel_stores]
            if show_backorder_only:
                rows = [r for r in rows if r.get("backorder_units", 0) > 0]

            disp = pd.DataFrame([{
                "Flag":          "⚠️ Backorder" if r.get("backorder_units", 0) > 0 else "",
                "SKU":           r["sku"],
                "Product":       r["product"],
                "Store":         r["to_store"],
                "Current SOH":   r["current_soh"],
                "Backorder Qty": r.get("backorder_units", 0),
                "Weekly Rate":   r["weekly_rate"],
                "Weeks Cover":   r["weeks_cover"],
                "Suggested Qty": r["suggested_qty"],
                "WH Available":  r["warehouse_available"],
            } for r in rows])
            st.dataframe(disp, width="stretch", hide_index=True, height=500)
            st.caption(
                f"{len(rows)} suggestion(s) shown · Target cover: {rec_data['thresholds']['replen_target_weeks']} weeks · "
                f"Negative SOH = pre-orders/backorders already owed to customers."
            )

    with sub_transfer:
        st.caption(
            "One store has slow-moving excess of a SKU while another store sells "
            "it well (or already owes pre-orders) but is low on stock — suggests "
            "moving stock between stores instead of ordering more."
        )
        if not transfers:
            st.success("✅ No transfer opportunities detected right now.")
        else:
            urgent_n = sum(1 for t in transfers if t.get("urgent_backorder"))
            if urgent_n:
                st.warning(f"⚠️ {urgent_n} of these would resolve an existing backorder/pre-order.")

            disp = pd.DataFrame([{
                "Flag":           "⚠️ Backorder" if t.get("urgent_backorder") else "",
                "SKU":            t["sku"],
                "Product":        t["product"],
                "Move From":      t["from_store"],
                "From Store SOH": t["from_soh"],
                "From Store Cover (wks)": t["from_weeks_cover"],
                "Move To":        t["to_store"],
                "To Store Backorder": t.get("to_backorder_units", 0),
                "To Store Sell Rate/wk": t["to_weekly_rate"],
                "Suggested Qty":  t["suggested_qty"],
            } for t in transfers])
            st.dataframe(disp, width="stretch", hide_index=True, height=500)
            st.caption(
                f"{len(transfers)} transfer suggestion(s) shown · "
                f"\"Move From\" = store with slow excess stock · \"Move To\" = store that needs it"
            )

    with sub_clearance:
        st.caption(
            "Excess stock everywhere with minimal recent sales — a transfer "
            "won't fix this, consider a markdown or clearance promotion instead."
        )
        if not clearance:
            st.success("✅ No clearance candidates flagged.")
        else:
            disp = pd.DataFrame([{
                "SKU":            c["sku"],
                "Product":        c["product"],
                "Total Units":    c["total_units"],
                "Weeks Cover":    c["weeks_cover"],
                "L6W Sold":       c["l6w_units_sold"],
                "Inventory Cost": f"${c['inventory_cost']:,.0f}",
            } for c in clearance])
            st.dataframe(disp, width="stretch", hide_index=True, height=500)
            total_tied_up = sum(c["inventory_cost"] for c in clearance)
            st.caption(f"{len(clearance)} candidate(s) shown · ${total_tied_up:,.0f} total inventory cost tied up")


def show_shipment_hub():
    st.subheader("📦 Shipment Hub")

    db_path  = os.path.join(DATA_DIR, "shipment_db.json")
    mtime    = os.path.getmtime(db_path) if os.path.exists(db_path) else 0
    shipments = load_shipment_db(_mtime=mtime)

    if not shipments:
        st.warning("No shipment data found. Make sure shipment_db.json is in the data folder.")
        return

    STATUS_COLORS = {
        "Delivered":            "🟢",
        "Arrived NZ":           "🔵",
        "Customs Cleared":      "🔵",
        "In Customs":           "🟡",
        "Documents Received":   "🟡",
        "In Transit":           "🟠",
        "Waiting for Shipping": "🔘",
        "Pending":              "⚪",
    }

    STATUS_ORDER = [
        "In Transit","Arrived NZ","Waiting for Shipping","Delivered",
        "Documents Received","In Customs","Customs Cleared","Pending"
    ]

    # Simple display options (same as Overview)
    STATUS_DISPLAY = ["In Transit", "Arrived NZ", "Waiting for Shipping", "Delivered"]

    ACTIVE_STATUS = {
        "Pending","Waiting for Shipping","In Transit",
        "Documents Received","In Customs","Arrived NZ","Customs Cleared"
    }

    # ── Build master dataframe ────────────────────────────────
    rows = []
    for inv_no, shp in shipments.items():
        raw_status = shp.get("tnl_status", "Pending")
        # Merge Pending, Documents Received, In Customs, Customs Cleared into simpler display
        if raw_status in ("Pending",):
            display_status = "Waiting for Shipping"
        elif raw_status in ("Documents Received", "In Customs", "Customs Cleared"):
            display_status = "Arrived NZ"
        else:
            display_status = raw_status
        rows.append({
            "Invoice No":      inv_no,
            "Store":           shp.get("store", ""),
            "Date":            shp.get("recv_date", ""),
            "Customer Ref":    shp.get("customer_ref", ""),
            "Shipping":        shp.get("shipping", ""),
            "Status":          display_status,
            "ETA":             shp.get("eta", ""),
            "EDD":             shp.get("edd", ""),
            "Actual Delivery": shp.get("actual_delivery", ""),
            "SKUs":            len(shp.get("items", [])),
            "Units":           sum(i.get("quantity", 0) for i in shp.get("items", [])),
            "Value (AUD)":     sum(i.get("exc_value", 0) for i in shp.get("items", [])),
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["_sort"] = df["Status"].apply(
            lambda x: STATUS_DISPLAY.index(x) if x in STATUS_DISPLAY else 99)
        df = df.sort_values(["_sort","ETA","Date"]).drop(columns=["_sort"])

    # ── Summary metrics ───────────────────────────────────────
    active_df    = df[df["Status"].isin(["In Transit","Arrived NZ","Waiting for Shipping"])]
    waiting      = len(df[df["Status"] == "Waiting for Shipping"])
    in_transit   = len(df[df["Status"] == "In Transit"])
    in_nz        = len(df[df["Status"] == "Arrived NZ"])
    delivered    = len(df[df["Status"] == "Delivered"])
    active_val   = active_df["Value (AUD)"].sum()
    active_units = int(active_df["Units"].sum())

    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("🔘 Waiting",    waiting)
    m2.metric("🟠 In Transit", in_transit)
    m3.metric("🟡 In NZ",      in_nz)
    m4.metric("🟢 Delivered",  delivered)
    m5.metric("📦 Active Units", f"{active_units:,}")
    m6.metric("💰 Active Value", f"${active_val:,.0f}")

    st.divider()

    # ── Load supporting data ──────────────────────────────────
    po_mtime  = _get_mtime(os.path.join(DATA_DIR, "preorders.json"))
    preorders = load_preorders(_mtime=po_mtime)

    co_df_hub = load_company(_mtime=_get_mtime(os.path.join(DATA_DIR,"output_company.parquet")))
    sku_cat_map = dict(zip(
        co_df_hub["supplier_code"].astype(str).str.upper(),
        co_df_hub["category"].fillna("Unknown")
    ))

    tab_ov, tab_detail, tab_search, tab_po, tab_pending, tab_limits, tab_manual = st.tabs([
        "📋 Overview", "🚢 Shipment Details", "🔍 Product Search",
        "🛒 Pre-orders", "⏳ Pending Orders (Not Invoiced)", "📦 Store Limits", "✏️ Manual Update"
    ])

    # ════════════════════════════════════════════════════════
    # Tab 1: Overview
    # ════════════════════════════════════════════════════════
    with tab_ov:
        # Filters
        fc1, fc2, fc3 = st.columns(3)
        status_opts    = [s for s in STATUS_DISPLAY if s in df["Status"].values]
        default_status = [s for s in status_opts if s != "Delivered"]
        sel_status   = fc1.multiselect("Status",   status_opts, default=default_status, key="ov_status")
        sel_store    = fc2.multiselect("Store",    sorted(df["Store"].unique()), key="ov_store")
        sel_shipping = fc3.multiselect("Shipping", sorted(df["Shipping"].dropna().unique()), key="ov_shipping")

        fdf = df.copy()
        if sel_status:   fdf = fdf[fdf["Status"].isin(sel_status)]
        if sel_store:    fdf = fdf[fdf["Store"].isin(sel_store)]
        if sel_shipping: fdf = fdf[fdf["Shipping"].isin(sel_shipping)]

        # Display table
        disp = fdf.copy()
        disp["Status"] = disp["Status"].apply(lambda x: f"{STATUS_COLORS.get(x,'⚪')} {x}")
        disp["Value (AUD)"] = disp["Value (AUD)"].apply(lambda x: f"${x:,.2f}" if x else "—")
        st.dataframe(
            disp[["Invoice No","Store","Date","Customer Ref","Shipping",
                  "Status","ETA","EDD","Actual Delivery","SKUs","Units","Value (AUD)"]],
            width="stretch", hide_index=True, height=400
        )
        st.caption(f"Showing {len(fdf)} of {len(df)} shipments")

        # SKU Summary download
        st.markdown("---")
        if st.button("📥 Download SKU Shipment Summary", key="sku_sum_prep"):
            with st.spinner("Generating..."):
                from collections import defaultdict
                ALL_STATUSES = ["Delivered","Arrived NZ","Customs Cleared",
                                "In Customs","Documents Received","In Transit","Pending"]
                STORE_ORDER  = ["Dunedin","Papanui","Queensgate","Riccarton",
                                "Richmond","Sylvia Park","Te Rapa","Office","Warehouse AKL"]
                sku_data = defaultdict(lambda: {
                    "Description": "", "Total": 0,
                    **{s: 0 for s in ALL_STATUSES},
                    **{stn: 0 for stn in STORE_ORDER},
                })
                for inv_no, shp in shipments.items():
                    status = shp.get("tnl_status","Pending")
                    store  = shp.get("store","")
                    for item in shp.get("items",[]):
                        sku  = item.get("sku","").strip()
                        qty  = item.get("quantity",0) or 0
                        if not sku: continue
                        sku_data[sku]["Description"] = item.get("description","")
                        sku_data[sku]["Total"]       += qty
                        if status in ALL_STATUSES: sku_data[sku][status] += qty
                        if store  in STORE_ORDER:  sku_data[sku][store]  += qty

                rows_sum = []
                for sku, d in sorted(sku_data.items()):
                    row = {"SKU": sku, "Description": d["Description"], "Total": d["Total"]}
                    for s in ALL_STATUSES: row[s] = d[s] if d[s] > 0 else ""
                    for stn in STORE_ORDER: row[stn] = d[stn] if d[stn] > 0 else ""
                    rows_sum.append(row)

                sum_df = pd.DataFrame(rows_sum)
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    sum_df.to_excel(writer, index=False, sheet_name="SKU Summary")
                    ws_s = writer.sheets["SKU Summary"]
                    hdr_fill = PatternFill("solid", fgColor="37474F")
                    hdr_font = Font(color="FFFFFF", bold=True)
                    status_colors_xl = {
                        "Delivered":"C8E6C9","Arrived NZ":"B3E5FC",
                        "Customs Cleared":"B3E5FC","In Customs":"FFF9C4",
                        "Documents Received":"FFF9C4","In Transit":"FFE0B2","Pending":"F5F5F5",
                    }
                    for ci, col in enumerate(sum_df.columns, 1):
                        cell = ws_s.cell(row=1, column=ci)
                        cell.font = Font(color="FFFFFF", bold=True) if col not in status_colors_xl else Font(bold=True)
                        cell.fill = PatternFill("solid", fgColor=status_colors_xl.get(col,"37474F"))
                    ws_s.column_dimensions["A"].width = 14
                    ws_s.column_dimensions["B"].width = 42
                    ws_s.column_dimensions["C"].width = 8
                    for i in range(4, len(sum_df.columns)+1):
                        ws_s.column_dimensions[get_column_letter(i)].width = 13
                    ws_s.freeze_panes = "A2"

                st.download_button("⬇️ Download SKU Summary", buf.getvalue(),
                    "PopStop_Shipment_SKU_Summary.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="sku_sum_dl")

    # ════════════════════════════════════════════════════════
    # Tab 2: Shipment Details (grouped by Customer Ref)
    # ════════════════════════════════════════════════════════
    with tab_detail:
        from collections import defaultdict

        # Filter controls — same 4 options as Overview
        fd1, fd2 = st.columns(2)
        det_status = fd1.multiselect("Filter by Status", STATUS_DISPLAY,
                         default=[s for s in STATUS_DISPLAY if s != "Delivered"],
                         key="det_status")
        det_store  = fd2.multiselect("Filter by Store",
                         sorted(df["Store"].unique()), key="det_store")

        fdf2 = df.copy()
        if det_status: fdf2 = fdf2[fdf2["Status"].isin(det_status)]
        if det_store:  fdf2 = fdf2[fdf2["Store"].isin(det_store)]

        if fdf2.empty:
            st.info("No shipments match the selected filters.")
        else:
            # Group by Customer Ref (shipment batch)
            ref_groups = defaultdict(list)
            for _, row in fdf2.iterrows():
                ref = row["Customer Ref"] or "Waiting for Shipping"
                ref_groups[ref].append(row)

            def sort_ref(ref):
                rows_in_ref = ref_groups[ref]
                status = rows_in_ref[0]["Status"]
                eta    = rows_in_ref[0]["ETA"] or "9999"
                return (STATUS_ORDER.index(status) if status in STATUS_ORDER else 99, eta)

            sorted_refs = sorted(ref_groups.keys(), key=sort_ref)

            for ref in sorted_refs:
                group_rows = ref_groups[ref]

                # Aggregate across all stores in this batch
                grp_sku_qty  = defaultdict(int)
                grp_sku_desc = {}
                grp_sku_exc  = defaultdict(float)
                grp_cat_skus = defaultdict(set)
                grp_cat_qty  = defaultdict(int)
                grp_status   = group_rows[0]["Status"]
                grp_eta      = min((r["ETA"] for r in group_rows if r["ETA"]), default="")
                grp_edd      = min((r["EDD"] for r in group_rows if r["EDD"]), default="")

                for row in group_rows:
                    for item in shipments.get(row["Invoice No"], {}).get("items", []):
                        sku  = item.get("sku","").strip()
                        qty  = item.get("quantity",0) or 0
                        cat  = sku_cat_map.get(sku.upper(), "Unknown")
                        grp_sku_qty[sku]  += qty
                        grp_sku_desc[sku]  = item.get("description","")
                        grp_sku_exc[sku]  += item.get("exc_value",0)
                        grp_cat_skus[cat].add(sku)
                        grp_cat_qty[cat]  += qty

                icon      = STATUS_COLORS.get(grp_status, "⚪")
                n_stores  = len(group_rows)
                n_skus    = len(grp_sku_qty)
                tot_units = sum(grp_sku_qty.values())
                tot_val   = sum(grp_sku_exc.values())
                shipping  = group_rows[0]["Shipping"] or "—"
                eta_label = f" | ETA {grp_eta}" if grp_eta else ""
                edd_label = f" | EDD {grp_edd}" if grp_edd else ""

                expander_label = (
                    f"{icon} **{ref}** — {shipping} | "
                    f"{n_stores} stores | {n_skus} SKUs | "
                    f"{tot_units:,} units | ${tot_val:,.0f}{eta_label}{edd_label}"
                )

                with st.expander(expander_label, expanded=False):

                    # Store summary table at the top
                    store_summary = []
                    for row in sorted(group_rows, key=lambda r: r["Store"]):
                        inv_items = shipments.get(row["Invoice No"], {}).get("items", [])
                        store_summary.append({
                            "Store":     row["Store"],
                            "Invoice":   row["Invoice No"],
                            "Status":    f"{STATUS_COLORS.get(row['Status'],'⚪')} {row['Status']}",
                            "ETA":       row["ETA"] or "—",
                            "EDD":       row["EDD"] or "—",
                            "Delivered": row["Actual Delivery"] or "—",
                            "SKUs":      len(inv_items),
                            "Units":     sum(i.get("quantity",0) for i in inv_items),
                            "Value":     f"${sum(i.get('exc_value',0) for i in inv_items):,.0f}",
                        })
                    st.dataframe(pd.DataFrame(store_summary),
                                 width="stretch", hide_index=True,
                                 height=min(55 + len(store_summary)*38, 320))

                    # Category breakdown
                    if grp_cat_qty:
                        cat_rows = sorted(grp_cat_qty.items(), key=lambda x: -x[1])
                        cat_df = pd.DataFrame([
                            {"Category": c, "SKUs": len(grp_cat_skus[c]), "Units": q}
                            for c, q in cat_rows
                        ])
                        st.markdown("**Category Breakdown:**")
                        st.dataframe(cat_df, width="stretch", hide_index=True,
                                     height=min(55 + len(cat_rows)*38, 280))

                    # Full product list
                    if grp_sku_qty:
                        prod_rows = sorted(grp_sku_qty.items(), key=lambda x: -x[1])
                        prod_df = pd.DataFrame([
                            {
                                "SKU":         sku,
                                "Description": grp_sku_desc.get(sku,""),
                                "Category":    sku_cat_map.get(sku.upper(),"Unknown"),
                                "Total Qty":   qty,
                                "Value (AUD)": f"${grp_sku_exc.get(sku,0):,.2f}",
                            }
                            for sku, qty in prod_rows
                        ])
                        st.markdown("**All Products (All Stores Combined):**")
                        st.dataframe(prod_df, width="stretch", hide_index=True,
                                     height=min(55 + len(prod_rows)*38, 450))

                    # ── Per-batch Excel download ──────────────────────
                    safe_ref = ref.replace(" ","_").replace("/","_")
                    dl_key = f"dl_batch_{safe_ref}"
                    if st.button(f"📥 Download {ref} Excel", key=dl_key):
                        with st.spinner("Generating..."):
                            buf = BytesIO()
                            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                                hdr_fill_dark = PatternFill("solid", fgColor="37474F")
                                hdr_fill_blue = PatternFill("solid", fgColor="1565C0")
                                hdr_font_w    = Font(color="FFFFFF", bold=True, size=9)
                                border        = _get_border()

                                def _write_batch_sheet(ws_b, rows_data, sheet_title):
                                    cols = ["SKU","Description","Category","Qty","Value (AUD)"]
                                    ws_b.row_dimensions[1].height = 22
                                    # Title row
                                    ws_b.cell(1,1, sheet_title).font = Font(bold=True, size=11, color="1565C0")
                                    ws_b.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols))
                                    # Header row
                                    ws_b.row_dimensions[2].height = 18
                                    for ci, h in enumerate(cols, 1):
                                        c = ws_b.cell(2, ci, h)
                                        c.fill   = hdr_fill_dark
                                        c.font   = hdr_font_w
                                        c.border = border
                                    # Data
                                    for ri, r in enumerate(rows_data, 3):
                                        fill = PatternFill("solid", fgColor="F5F5F5" if ri%2==0 else "FFFFFF")
                                        ws_b.row_dimensions[ri].height = 17
                                        for ci, v in enumerate(r, 1):
                                            c = ws_b.cell(ri, ci, v)
                                            c.fill = fill; c.border = border
                                            c.font = Font(size=9)
                                    # Widths
                                    ws_b.column_dimensions["A"].width = 14
                                    ws_b.column_dimensions["B"].width = 48
                                    ws_b.column_dimensions["C"].width = 18
                                    ws_b.column_dimensions["D"].width = 8
                                    ws_b.column_dimensions["E"].width = 14
                                    ws_b.freeze_panes = "A3"

                                # Sheet 1: All stores combined
                                all_rows_data = [
                                    (sku,
                                     grp_sku_desc.get(sku,""),
                                     sku_cat_map.get(sku.upper(),"Unknown"),
                                     qty,
                                     round(grp_sku_exc.get(sku,0),2))
                                    for sku, qty in sorted(grp_sku_qty.items(), key=lambda x:-x[1])
                                ]
                                all_df_xl = pd.DataFrame(all_rows_data,
                                    columns=["SKU","Description","Category","Qty","Value (AUD)"])
                                all_df_xl.to_excel(writer, index=False,
                                                   sheet_name="All Stores", startrow=1)
                                ws_all = writer.sheets["All Stores"]
                                title_str = (f"{ref} | {grp_status} | "
                                             f"ETA {grp_eta or 'TBC'} | EDD {grp_edd or 'TBC'} | "
                                             f"{n_skus} SKUs | {tot_units:,} units")
                                _write_batch_sheet(ws_all, all_rows_data, title_str)

                                # One sheet per store
                                for row_s in sorted(group_rows, key=lambda r: r["Store"]):
                                    store_name = row_s["Store"]
                                    inv_items  = shipments.get(row_s["Invoice No"],{}).get("items",[])
                                    if not inv_items:
                                        continue
                                    store_rows_data = sorted([
                                        (i.get("sku","").strip(),
                                         i.get("description",""),
                                         sku_cat_map.get(i.get("sku","").strip().upper(),"Unknown"),
                                         i.get("quantity",0) or 0,
                                         round(i.get("exc_value",0),2))
                                        for i in inv_items if i.get("quantity",0)
                                    ], key=lambda x: -x[3])
                                    if not store_rows_data:
                                        continue
                                    sheet_nm = store_name[:28]
                                    # avoid duplicate sheet names
                                    existing = [s.title for s in writer.book.worksheets]
                                    if sheet_nm in existing:
                                        sheet_nm = sheet_nm[:25] + "_2"
                                    store_df_xl = pd.DataFrame(store_rows_data,
                                        columns=["SKU","Description","Category","Qty","Value (AUD)"])
                                    store_df_xl.to_excel(writer, index=False,
                                                         sheet_name=sheet_nm, startrow=1)
                                    ws_st = writer.sheets[sheet_nm]
                                    store_units = sum(x[3] for x in store_rows_data)
                                    store_val   = sum(x[4] for x in store_rows_data)
                                    store_title = (f"{store_name} | {ref} | "
                                                   f"EDD {row_s['EDD'] or 'TBC'} | "
                                                   f"{len(store_rows_data)} SKUs | "
                                                   f"{store_units:,} units | ${store_val:,.0f}")
                                    _write_batch_sheet(ws_st, store_rows_data, store_title)

                        fname = f"Shipment_{safe_ref}_{grp_eta or 'TBC'}.xlsx"
                        st.download_button(
                            f"⬇️ Download {ref}",
                            buf.getvalue(),
                            file_name=fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_batch_btn_{safe_ref}",
                        )

    # ════════════════════════════════════════════════════════
    # Tab 3: Product Search
    # ════════════════════════════════════════════════════════
    with tab_search:
        st.markdown("Search by SKU or product name across all shipments and pre-orders.")
        term = st.text_input("Search", placeholder="e.g. FUN86294 or Spider-Man", key="hub_search")
        if term:
            results = []
            for inv_no, shp in shipments.items():
                for item in shp.get("items", []):
                    if (term.lower() in item.get("sku","").lower() or
                            term.lower() in item.get("description","").lower()):
                        results.append({
                            "Invoice No":     inv_no,
                            "Store":          shp.get("store",""),
                            "SKU":            item.get("sku",""),
                            "Description":    item.get("description",""),
                            "Qty":            item.get("quantity",0),
                            "Status":         f"{STATUS_COLORS.get(shp.get('tnl_status','Pending'),'⚪')} {shp.get('tnl_status','Pending')}",
                            "Customer Ref":   shp.get("customer_ref",""),
                            "ETA":            shp.get("eta",""),
                            "EDD":            shp.get("edd",""),
                            "Actual Delivery": shp.get("actual_delivery",""),
                        })

            po_results = []
            for sku_code, po in preorders.items():
                if (term.lower() in sku_code.lower() or
                        term.lower() in po.get("description","").lower()):
                    if "orders" in po and po["orders"]:
                        for order in po["orders"]:
                            po_results.append({
                                "SKU":           sku_code,
                                "Description":   po.get("description",""),
                                "Order No":      order.get("order_no",""),
                                "ETA":           order.get("eta",""),
                                "Pre-order Qty": order.get("qty",0),
                            })
                    else:
                        po_results.append({
                            "SKU":           sku_code,
                            "Description":   po.get("description",""),
                            "Order No":      po.get("order_no",""),
                            "ETA":           po.get("eta",""),
                            "Pre-order Qty": po.get("total_qty",0),
                        })

            pending_results = []
            pending_path = os.path.join(DATA_DIR, "pending_orders.json")
            if os.path.exists(pending_path):
                with open(pending_path) as f:
                    pending_data_search = json.load(f)
                search_sku_lookup = dict(zip(
                    co_df_hub["supplier_code"].astype(str).str.upper(),
                    co_df_hub["product"]
                )) if "product" in co_df_hub.columns else {}
                for p in pending_data_search.get("pending", []):
                    p_sku  = p.get("sku","")
                    p_desc = p.get("order_description","") or search_sku_lookup.get(p_sku,"")
                    if term.lower() in p_sku.lower() or term.lower() in p_desc.lower():
                        pending_results.append({
                            "Flag":           "🔴 Overdue" if p.get("overdue") else "🟡 Pending",
                            "SKU":            p_sku,
                            "Description":    p_desc,
                            "Ordered":        p.get("ordered_qty",0),
                            "Invoiced":       p.get("invoiced_qty",0),
                            "Remaining":      p.get("remaining_qty",0),
                            "Order Date":     p.get("order_date",""),
                            "Days Pending":   p.get("days_pending",0),
                        })

            if results:
                st.success(f"🚢 {len(results)} shipment line(s) found")
                st.dataframe(pd.DataFrame(results), width="stretch", hide_index=True)
            else:
                st.info("No shipments found.")

            if pending_results:
                st.success(f"⏳ {len(pending_results)} pending order line(s) found (ordered, not yet invoiced)")
                pend_df = pd.DataFrame(pending_results)
                st.metric("Total Remaining Qty", f"{int(pend_df['Remaining'].sum()):,} units")
                st.dataframe(pend_df, width="stretch", hide_index=True)
            else:
                st.info("No pending orders found.")

            if po_results:
                st.success(f"🛒 {len(po_results)} pre-order line(s) found")
                po_df = pd.DataFrame(po_results)
                qty_col = next((c for c in po_df.columns if "qty" in c.lower()), None)
                if qty_col:
                    st.metric("Total Pre-order Qty", f"{int(po_df[qty_col].sum()):,} units")
                st.dataframe(po_df, width="stretch", hide_index=True)
            else:
                st.info("No pre-orders found.")

    # ════════════════════════════════════════════════════════
    # Tab 4: Pre-orders
    # ════════════════════════════════════════════════════════
    with tab_po:
        st.markdown("### 🛒 Outstanding Pre-orders from IKON")
        if not preorders:
            st.warning("No pre-order data. Run `ikon_scraper_v5.py` to fetch from IKON.")
        else:
            po_rows = []
            for sku_code, po in preorders.items():
                if "orders" in po and po["orders"]:
                    for order in po["orders"]:
                        po_rows.append({
                            "SKU":         sku_code,
                            "Description": po.get("description",""),
                            "Order No":    order.get("order_no",""),
                            "Date":        order.get("date",""),
                            "ETA":         order.get("eta",""),
                            "Qty":         order.get("qty",0),
                        })
                else:
                    po_rows.append({
                        "SKU":         sku_code,
                        "Description": po.get("description",""),
                        "Order No":    po.get("order_no",""),
                        "Date":        "",
                        "ETA":         po.get("eta",""),
                        "Qty":         po.get("total_qty",0),
                    })

            po_df = pd.DataFrame(po_rows) if po_rows else pd.DataFrame(
                columns=["SKU","Description","Order No","Date","ETA","Qty"])

            c1, c2, c3 = st.columns(3)
            c1.metric("Total SKUs",   len(preorders))
            c2.metric("Total Orders", len(po_df))
            c3.metric("Total Units",  f"{int(po_df['Qty'].sum()):,}" if not po_df.empty else "0")

            search_po = st.text_input("🔍 Search SKU / Description", key="po_search")
            if search_po and not po_df.empty:
                po_df = po_df[
                    po_df["SKU"].astype(str).str.contains(search_po, case=False, na=False) |
                    po_df["Description"].astype(str).str.contains(search_po, case=False, na=False)
                ]
            if not po_df.empty:
                po_df = po_df.sort_values("Qty", ascending=False)

            st.dataframe(po_df, width="stretch", hide_index=True, height=500)
            st.caption(f"{len(preorders)} SKUs | Updated by IKON scraper")

    # ════════════════════════════════════════════════════════
    # Tab: Pending Orders (ordered but not yet invoiced)
    # ════════════════════════════════════════════════════════
    with tab_pending:
        st.markdown("### ⏳ Pending Orders — Ordered but Not Yet Invoiced")
        st.caption(
            "Tracks SKUs from your weekly order emails that haven't shown up "
            "(in sufficient quantity) on any invoice received after the order was sent. "
            "Prevents accidental duplicate ordering during the order→invoice gap."
        )

        pending_path = os.path.join(DATA_DIR, "pending_orders.json")
        if not os.path.exists(pending_path):
            st.info(
                "No pending_orders.json found yet. Run `track_pending_orders.py` "
                "on your local machine to generate this data."
            )
        else:
            with open(pending_path) as f:
                pending_data = json.load(f)

            pending_list = pending_data.get("pending", [])
            gen_at       = pending_data.get("generated_at", "")
            overdue_thr  = pending_data.get("overdue_days_threshold", 7)

            if gen_at:
                try:
                    gen_dt = datetime.strptime(gen_at[:19], "%Y-%m-%dT%H:%M:%S")
                    st.caption(f"📅 Last checked: {gen_dt.strftime('%d %b %Y %H:%M')}")
                except Exception:
                    pass

            if not pending_list:
                st.success("✅ No pending orders! Everything has been invoiced.")
            else:
                overdue_n = sum(1 for p in pending_list if p.get("overdue"))
                c1, c2, c3 = st.columns(3)
                c1.metric("Pending SKUs", len(pending_list))
                c2.metric(f"Overdue (>{overdue_thr}d)", overdue_n)
                c3.metric("Total Units Pending", sum(p.get("remaining_qty",0) for p in pending_list))

                st.markdown("---")

                show_overdue_only = st.checkbox("🔴 Show overdue only", key="pending_overdue_only")
                rows = [p for p in pending_list if (not show_overdue_only or p.get("overdue"))]

                sku_lookup = dict(zip(
                    co_df_hub["supplier_code"].astype(str).str.upper(),
                    co_df_hub["product"]
                )) if "product" in co_df_hub.columns else {}

                show_new_only = st.checkbox("🆕 Show new/unrecognized SKUs only", key="pending_new_only")

                disp_rows = []
                for p in rows:
                    in_system = p["sku"] in sku_lookup
                    if show_new_only and in_system:
                        continue

                    if in_system:
                        product_name = sku_lookup[p["sku"]]
                        flag = "🔴 Overdue" if p.get("overdue") else "🟡 Pending"
                    else:
                        # Not yet in output_company.parquet — fall back to the name
                        # from the order spreadsheet itself, and flag it so it's
                        # obvious this is a brand-new / not-yet-onboarded product.
                        product_name = p.get("order_description") or "(no name available)"
                        flag = "🆕 New SKU" + (" · 🔴 Overdue" if p.get("overdue") else "")

                    disp_rows.append({
                        "Flag":         flag,
                        "SKU":          p["sku"],
                        "Product":      product_name,
                        "Ordered":      p["ordered_qty"],
                        "✈️ Air":       p.get("ordered_air", 0),
                        "🚢 Sea":       p.get("ordered_sea", 0),
                        "Invoiced":     p["invoiced_qty"],
                        "Remaining":    p["remaining_qty"],
                        "Remaining Air": p.get("remaining_air", 0),
                        "Remaining Sea": p.get("remaining_sea", 0),
                        "Order Date":   p["order_date"],
                        "Days Pending": p["days_pending"],
                        "Order Email":  p["order_subject"],
                    })

                new_sku_count = sum(1 for p in rows if p["sku"] not in sku_lookup)
                if new_sku_count > 0:
                    st.info(
                        f"🆕 {new_sku_count} SKU(s) in this list aren't in the system yet "
                        f"(no sales history / not yet received) — their product names are "
                        f"taken directly from the order spreadsheet."
                    )

                pending_df = pd.DataFrame(disp_rows)
                st.dataframe(
                    pending_df, width="stretch", hide_index=True, height=500,
                    column_config={
                        "Days Pending": st.column_config.NumberColumn(format="%d days"),
                    }
                )
                st.caption(
                    f"{len(rows)} SKU(s) shown · Highlighted as overdue if pending "
                    f"more than {overdue_thr} days since order was sent."
                )

    # ════════════════════════════════════════════════════════
    # Tab 5: Store Limits
    # ════════════════════════════════════════════════════════
    with tab_limits:
        st.markdown("### 📦 Store Limits Check — Incoming Shipments")
        st.caption("See if incoming shipments will push categories over their store limits.")

        limits_df     = load_limits(_mtime=_get_mtime(LIMITS_FILE))
        all_stores_sl = sorted(df["Store"].unique().tolist())
        sel_store_sl  = st.selectbox("Select Store", all_stores_sl, key="sl_store")

        if sel_store_sl:
            from collections import defaultdict
            store_abbrev_map = {
                "Dunedin":"DUN","Papanui":"PAP","Queensgate":"QG","Riccarton":"RICC",
                "Richmond":"RICH","Sylvia Park":"SP","Te Rapa":"TR","Office":"Office",
            }
            soh_col = store_abbrev_map.get(sel_store_sl)

            # ── Gather all active shipments for this store ────────
            # Group by Customer Ref (shipment batch) to find next arrival
            batch_data = defaultdict(lambda: {
                "inv_nos": [], "edd": "", "eta": "", "customer_ref": "",
                "cat_skus": defaultdict(set), "cat_qty": defaultdict(int)
            })

            for inv_no, shp in shipments.items():
                if shp.get("store") != sel_store_sl: continue
                if shp.get("tnl_status") not in ACTIVE_STATUS: continue
                ref = shp.get("customer_ref") or shp.get("tnl_job_no") or inv_no
                edd = shp.get("edd","") or shp.get("eta","")
                batch_data[ref]["inv_nos"].append(inv_no)
                batch_data[ref]["customer_ref"] = ref
                if not batch_data[ref]["edd"] or (edd and edd < batch_data[ref]["edd"]):
                    batch_data[ref]["edd"] = edd
                    batch_data[ref]["eta"] = shp.get("eta","")
                for item in shp.get("items", []):
                    sku = item.get("sku","").strip().upper()
                    qty = item.get("quantity",0) or 0
                    cat = sku_cat_map.get(sku, "Unknown")
                    batch_data[ref]["cat_skus"][cat].add(sku)
                    batch_data[ref]["cat_qty"][cat] += qty

            if not batch_data:
                st.info(f"No active incoming shipments for {sel_store_sl}.")
            else:
                # Find next shipment (earliest EDD)
                sorted_batches = sorted(
                    batch_data.items(),
                    key=lambda x: x[1]["edd"] or "9999"
                )
                next_ref, next_batch = sorted_batches[0]

                cur_soh = {}
                if soh_col and soh_col in co_df_hub.columns:
                    cur_soh = co_df_hub.groupby("category")[soh_col].sum().to_dict()

                store_limits = limits_df[limits_df["store"] == sel_store_sl] if not limits_df.empty else pd.DataFrame()

                def build_table(cat_qty, cat_skus, eta_override=None):
                    rows_out = []
                    for cat in sorted(cat_qty.keys()):
                        inc_units = cat_qty[cat]
                        inc_skus  = len(cat_skus[cat])
                        cur_units = int(cur_soh.get(cat, 0) or 0)
                        total     = cur_units + inc_units
                        lim_row   = store_limits[store_limits["category"] == cat] if not store_limits.empty else pd.DataFrame()
                        lim_units_raw = lim_row["max_units"].values[0] if not lim_row.empty else None
                        lim_skus_raw  = lim_row["max_sku"].values[0]   if not lim_row.empty else None
                        lim_units = int(lim_units_raw) if lim_units_raw is not None and pd.notna(lim_units_raw) else None
                        lim_skus  = int(lim_skus_raw)  if lim_skus_raw  is not None and pd.notna(lim_skus_raw)  else None
                        if lim_units:
                            over = total - lim_units
                            status_sl = f"⚠️ Over by {over}" if over > 0 else ("🟡 Near limit" if (lim_units - total) < lim_units * 0.1 else "✅ OK")
                        else:
                            status_sl = "— No limit set"
                        rows_out.append({
                            "Category":      cat,
                            "Current SOH":   cur_units,
                            "Incoming":      inc_units,
                            "Incoming SKUs": inc_skus,
                            "Total After":   total,
                            "Limit Units":   lim_units or "—",
                            "Limit SKUs":    lim_skus or "—",
                            "Status":        status_sl,
                        })
                    return pd.DataFrame(rows_out)

                # ── View toggle ───────────────────────────────────
                view_sl = st.radio("View", ["🚀 Next Shipment", "📦 All Incoming"],
                                   horizontal=True, key="sl_view")

                if view_sl == "🚀 Next Shipment":
                    edd_label = next_batch["edd"] or "TBC"
                    eta_label = next_batch["eta"] or "TBC"
                    n_skus    = sum(len(v) for v in next_batch["cat_skus"].values())
                    n_units   = sum(next_batch["cat_qty"].values())

                    st.markdown(f"**Next shipment: `{next_ref}`**")
                    nm1, nm2, nm3 = st.columns(3)
                    nm1.metric("ETA",   eta_label)
                    nm2.metric("EDD",   edd_label)
                    nm3.metric("Total Units", f"{n_units:,} ({n_skus} SKUs)")

                    next_df = build_table(next_batch["cat_qty"], next_batch["cat_skus"])
                    over_next = [r for _, r in next_df.iterrows() if "⚠️" in str(r["Status"])]
                    if over_next:
                        st.error(f"⚠️ {len(over_next)} categories will exceed limits")
                    else:
                        st.success("✅ All categories within limits")
                    st.dataframe(next_df, width="stretch", hide_index=True,
                                 height=min(55 + len(next_df)*38, 500))

                    # Other batches coming after
                    if len(sorted_batches) > 1:
                        st.markdown("---")
                        st.markdown(f"**Other upcoming batches for {sel_store_sl}:**")
                        other_rows = []
                        for ref, batch in sorted_batches[1:]:
                            other_rows.append({
                                "Customer Ref": ref,
                                "EDD":          batch["edd"] or "TBC",
                                "Categories":   len(batch["cat_qty"]),
                                "SKUs":         sum(len(v) for v in batch["cat_skus"].values()),
                                "Units":        sum(batch["cat_qty"].values()),
                            })
                        st.dataframe(pd.DataFrame(other_rows), width="stretch",
                                     hide_index=True)

                else:  # All Incoming
                    # Aggregate across all batches
                    all_cat_qty  = defaultdict(int)
                    all_cat_skus = defaultdict(set)
                    all_cat_edd  = {}
                    for ref, batch in sorted_batches:
                        for cat, qty in batch["cat_qty"].items():
                            all_cat_qty[cat]  += qty
                            all_cat_skus[cat] |= batch["cat_skus"][cat]
                            edd = batch["edd"]
                            if cat not in all_cat_edd or (edd and edd < all_cat_edd[cat]):
                                all_cat_edd[cat] = edd

                    all_df = build_table(all_cat_qty, all_cat_skus)
                    # Add EDD column
                    all_df["Next EDD"] = all_df["Category"].map(all_cat_edd).fillna("TBC")
                    # Reorder columns
                    cols = ["Category","Current SOH","Incoming","Incoming SKUs",
                            "Total After","Limit Units","Limit SKUs","Next EDD","Status"]
                    all_df = all_df[[c for c in cols if c in all_df.columns]]

                    over_all = [r for _, r in all_df.iterrows() if "⚠️" in str(r["Status"])]
                    if over_all:
                        st.error(f"⚠️ {len(over_all)} categories will exceed limits")
                    else:
                        st.success("✅ All categories within limits")
                    st.dataframe(all_df, width="stretch", hide_index=True,
                                 height=min(55 + len(all_df)*38, 500))
                    st.caption(f"{len(sorted_batches)} active shipment batch(es) for {sel_store_sl}")

    # ════════════════════════════════════════════════════════
    # Tab 6: Manual Update
    # ════════════════════════════════════════════════════════
    with tab_manual:
        st.markdown("### ✏️ Manual Update")
        st.caption("Update shipment status and dates manually.")
        selected = st.selectbox("Select Invoice", sorted(shipments.keys(), reverse=True), key="manual_inv")
        if selected:
            shp = shipments[selected]
            st.markdown(
                f"**Invoice #{selected}** | {shp.get('store','')} | "
                f"Customer Ref: `{shp.get('customer_ref','—')}`"
            )
            c1, c2 = st.columns(2)
            status_opts = ["Waiting for Shipping","Pending","In Transit",
                           "Documents Received","In Customs","Customs Cleared",
                           "Arrived NZ","Delivered"]
            cur_status   = shp.get("tnl_status","Pending")
            new_status   = c1.selectbox("Status", status_opts,
                               index=status_opts.index(cur_status) if cur_status in status_opts else 0)
            new_job      = c1.text_input("TNL Job No",       value=shp.get("tnl_job_no",""), key="manual_job")
            new_eta      = c2.text_input("ETA (YYYY-MM-DD)", value=shp.get("eta",""), key="manual_eta")
            new_edd      = c2.text_input("EDD (YYYY-MM-DD)", value=shp.get("edd",""), key="manual_edd")
            new_delivery = c2.text_input("Actual Delivery",  value=shp.get("actual_delivery",""), key="manual_del")

            if st.button("💾 Save Changes", key="manual_save"):
                full_path = os.path.join(DATA_DIR, "shipment_db.json")
                try:
                    with open(full_path) as f:
                        full_db = json.load(f)
                    if not isinstance(full_db, dict):
                        full_db = {"shipments": {}}
                    if "shipments" not in full_db:
                        full_db["shipments"] = {}
                    if selected not in full_db["shipments"]:
                        full_db["shipments"][selected] = dict(shp)
                    full_db["shipments"][selected].update({
                        "tnl_status":      new_status,
                        "tnl_job_no":      new_job,
                        "eta":             new_eta,
                        "edd":             new_edd,
                        "actual_delivery": new_delivery,
                        "manual_override": True,
                    })
                    full_db["last_updated"] = datetime.now().isoformat()
                    with open(full_path, "w") as f:
                        json.dump(full_db, f, indent=2)
                    st.success(f"✅ Invoice #{selected} updated!")
                    st.cache_data.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Save failed: {e}")

    if not shipments:
        st.warning("No shipment data found. Make sure shipment_db.json is in the data folder.")


# ============================================================
# 渲染
# ============================================================
if role in ["admin", "manager"]:
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "🏢 Company View",
        "🏪 Store View",
        "📊 Store Limits",
        "📋 Last Week Summary",
        "📈 Weekly Reports",
        "🛒 Store Requests",
        "📦 Shipment Hub",
        "🔄 Inventory Recommendations"
    ])
    with tab1:
        show_company_view()
    with tab2:
        show_store_view()
    with tab3:
        show_limits_view()
    with tab4:
        show_last_week_summary()
    with tab5:
        show_weekly_view()
    with tab6:
        show_store_requests()
    with tab7:
        show_shipment_hub()
    with tab8:
        show_inventory_recommendations()
else:
    tab1, tab2 = st.tabs(["🏪 Store View", "📊 Store Limits"])
    with tab1:
        show_store_view(store_filter=user_store)
    with tab2:
        show_limits_view(store_filter=user_store)

